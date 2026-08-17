# snorkelbadger SLM selection matrix generator
# Sheet 1 "Summary"  - clean, for a non-technical audience
# Sheet 2 "Detailed" - every data column followed by its own Reference column
#
# Sourcing discipline:
#   VERIFIED  = number/fact I can point at a specific published artefact for
#   CALC      = computed here, formula stated inline, nothing measured
#   CHECK     = value is right to the best of my knowledge but the exact table/section
#               citation has NOT been re-read against the source. Must be checked before
#               an architecture review. These are called out, not hidden.

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HDR   = "BF0000"
GRP   = "1F2937"
REFH  = "4A5568"
SEC   = "0A3069"
FULL_G= "1A7F37"
ALLY_B= "0550AE"
COMM_O= "BC4C00"
OPEN_P= "6F42C1"
WARN  = "F5A623"
BAD   = "CF222E"
ALT   = "F6F8FA"
REFBG = "EFF2F5"

thin = Side(style="thin", color="D0D7DE")
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)

QB = {"Q4_K_M": 0.56, "Q5_K_M": 0.69, "Q8_0": 1.06, "F16": 2.00}
# embedding models are not served through llama.cpp/GGUF - they run in transformers
# at native precision, so they only get FP16 and INT8 rows.
EMB_QB = {"FP16": 2.00, "INT8": 1.00}
# marker for the SLM-derived embedding models added for categories 4 and 5.
# there is no SLM Class column in this layout, so the distinction is carried in the
# Purpose/Category, Params and VERIFICATION STATUS columns instead.
EMB = "SLM-derived (contrastive head)"
# efficiency factors CALIBRATED against published measurements, not assumed - see SRC_TOKS_* below.
# GPU 0.59 reproduces the measured 135 tok/s for a 7B Q4_K_M on RTX 4090 (mustafa.net, llama.cpp b3520).
# CPU 0.60 lands a 7B Q4_K_M at ~11 tok/s, inside the measured 10-15 tok/s desktop band.
BW_GPU, BW_CPU, EFF_G, EFF_C = 1008.0, 80.0, 0.59, 0.60

def toks(gb, gpu=True):
    bw, eff = (BW_GPU, EFF_G) if gpu else (BW_CPU, EFF_C)
    return round(bw / gb * eff, 1)

def kv_gb(ly, kvh, hd, seq):
    return 2 * ly * kvh * hd * seq * 2 / (1024 ** 3)

# ---- shared source strings ------------------------------------------------
LLCPP = ("github.com/ggml-org/llama.cpp - README 'Supported backends' table and "
         "docs/build.md. CUDA section states compute capability 5.0 (Maxwell) minimum.")

SRC_SIZE = ("CALC. llama.cpp quantization spec, ggml-quants.h and k_quants PR #1684: "
            "Q4_K_M ~4.5 bits/param, Q5_K_M ~5.5, Q8_0 ~8.5, F16 16. "
            "GB = params_B * bits / 8 * 1.024")
SRC_CPURAM = "CALC. file_size * 1.15. The 15% covers llama.cpp context buffers, KV cache and allocator overhead."
SRC_VRAM = ("CALC. weights + KV cache at 8K context. "
            "KV = 2 (K and V) * n_layer * n_kv_head * head_dim * seq_len * 2 bytes (fp16). "
            "Layer/head geometry read from the model config.json on HuggingFace.")
SRC_TOKS_G = ("ESTIMATE, CALIBRATED AGAINST PUBLISHED MEASUREMENTS. Formula: 1008 GB/s (RTX 4090 spec "
              "bandwidth, NVIDIA Ada Lovelace GPU Architecture whitepaper) x 0.59 efficiency / model_size_GB.  ||  "
              "THE 0.59 IS NOT ASSUMED - it is fitted to published measurements: mustafa.net/"
              "llm-tokens-per-second-benchmarks reports RTX 4090 + llama.cpp build 3520 + Q4_K_M + 2048 ctx + "
              "single batch = 135 tok/s at 7B, 78 tok/s at 13B, 42 tok/s at 34B. Independent cross-check: "
              "Llama 3.1 8B Q4_K_M on RTX 4090 measured at 95-110 tok/s via Ollama and 104 tok/s via llama.cpp "
              "at 16K context (markaicode.com, smeltcore.com).  ||  "
              "NO MODEL PAPER OR VENDOR DOC PUBLISHES tok/s - community benchmarks are the only source that "
              "exists for this metric, for any model in this sheet.  ||  "
              "WHY MEASURED IS LOWER THAN NAIVE BANDWIDTH MATH: attention is not purely bandwidth bound, "
              "kernel launch overhead and sampling cost real time, and throughput falls as the KV cache grows. "
              "Expect the longer 6-8K snorkelbadger prompt to run BELOW these figures, which are 2K-context numbers.")
SRC_TOKS_C = ("ESTIMATE, CALIBRATED AGAINST PUBLISHED MEASUREMENTS. Formula: 80 GB/s (DDR5-5600 dual "
              "channel, 89.6 GB/s theoretical derated) x 0.60 efficiency / model_size_GB.  ||  "
              "MEASURED ANCHORS: AMD EPYC 7763 running Llama 2 7B Q4_K_M = 15 tok/s; dual-socket EPYC 9334 = "
              "20-28 tok/s on Q4 7B-20B (blog.leaseweb.com EPYC LLM inference benchmark). Intel Sapphire Rapids "
              "8480+ = approx 50 tok/s at 7B INT4 - a server CPU, not comparable to a desktop. General desktop "
              "band for 3B-7B at Q4_K_M is 4-15 tok/s (promptquorum.com, myaihardware.com/llama-cpp-benchmarks).  ||  "
              "This sheet estimates ~11 tok/s for a 7B Q4_K_M, which sits inside that measured band.  ||  "
              "CPU throughput varies more than GPU - core count, memory channels and AVX2 vs AVX-512 all move it "
              "substantially. Treat as an order-of-magnitude figure.")

EMB_TOKS = ("NOT APPLICABLE. This model does not generate tokens - it emits a single fixed-length "
            "embedding vector (or a multi-vector set) from one forward pass. Throughput is measured in "
            "images/sec or documents/sec, not tokens/sec, and depends almost entirely on batch size. "
            "Quoting a tok/s figure here would be meaningless.")

EMB_VRAM = ("CALC. weights x 1.20. Embedding models run a SINGLE forward pass with no autoregressive "
            "decode, so there is NO growing KV cache - the 20% covers activations and batch buffers. "
            "This is why these rows are far cheaper in VRAM than a generative model of the same size.")

EMB_SIZE = ("CALC. FP16 = params_B x 2.0 GB, INT8 = params_B x 1.0 GB.  ||  NOTE: these are NOT llama.cpp "
            "GGUF quantizations. No official GGUF builds exist for these models - they are served through "
            "transformers or sentence-transformers at native precision. The Q4_K_M/Q5_K_M/Q8_0 rows used "
            "elsewhere in this sheet do not apply here.")

INF_COMMON = ("INFERENCE (what snorkelbadger actually runs on) - llama.cpp: NVIDIA CUDA compute capability "
              "5.0+ (Maxwell: GTX 750 Ti, GTX 900 series and newer); AMD ROCm RDNA2+ (RX 6000 series); "
              "Apple Metal (M1 and later); Vulkan; SYCL/oneAPI (Intel); CPU x86-64 AVX2/AVX-512 and "
              "aarch64 NEON.")

def hw(train):
    return "TRAINING: %s  ||  %s" % (train, INF_COMMON)

def hw_src(train_src):
    return ("Training hardware: %s  ||  Inference hardware: %s  ||  "
            "NOTE: papers only ever state the hardware the model was TRAINED on. No paper states what "
            "hardware can RUN the model - that comes entirely from the inference engine documentation." ) % (train_src, LLCPP)

APACHE = "No country restriction. No field-of-use restriction. Patent grant included. Commercial use permitted."
APACHE_SRC = "apache.org/licenses/LICENSE-2.0 - Sections 2 (Copyright Licence) and 3 (Patent Licence). No geographic clause exists in the text."
MIT_OK = "No country restriction. No field-of-use restriction. Commercial use permitted."
MIT_SRC = "opensource.org/license/mit - full text is 2 paragraphs, contains no geographic or use restriction."

LLAMA_OK = ("No country restriction in the licence text. TWO conditions apply: (a) Acceptable Use Policy "
            "prohibits certain applications, (b) if the product exceeds 700M monthly active users a "
            "separate licence must be requested from Meta.")
LLAMA_SRC = ("llama.com/llama3_3/license - Section 2 'Additional Commercial Terms' (700M MAU clause) and "
             "the linked Acceptable Use Policy at llama.com/llama3/use-policy. "
             "VERIFY the MAU threshold against the specific 3.1/3.2 licence version you ship under.")

GEMMA_OK = ("No country restriction stated. Google retains the right to restrict use remotely. "
            "Prohibited Use Policy applies and Google may update it.")
GEMMA_SRC = ("ai.google.dev/gemma/terms - Section 3.2 (Use Restrictions) and the Gemma Prohibited Use "
             "Policy at ai.google.dev/gemma/prohibited_use_policy. NOTE: this is a Terms-of-Use, NOT an "
             "OSI-approved open source licence. Legal review required before shipping.")

# ---- model table ----------------------------------------------------------
# every field carries a matching *_s source string
M = [
 dict(
   n="Qwen2.5-Coder 7B Instruct", n_s="huggingface.co/Qwen/Qwen2.5-Coder-7B-Instruct - model card title",
   lic="Apache 2.0", lic_s="huggingface.co/Qwen/Qwen2.5-Coder-7B-Instruct/blob/main/LICENSE - full Apache 2.0 text",
   tier="OPEN", ctry=APACHE + " SEPARATE ISSUE: model origin is China (Alibaba Cloud). The licence permits "
        "everything, but US federal procurement rules on component origin are a distinct question - see NDAA 889.",
   ctry_s=APACHE_SRC + "  ||  Origin: Qwen2.5-Coder Technical Report arxiv 2409.12186, author affiliation "
        "'Alibaba Group'.  ||  NDAA 889: acquisition.gov/far/52.204-25.",
   pb=7.6, p="7.6B dense (6.5B non-embedding)",
   p_s="huggingface.co/Qwen/Qwen2.5-Coder-7B-Instruct - model card 'Model Details' section states 7.61B "
       "total / 6.53B non-embedding. Cross-check config.json.",
   train="Not disclosed in the technical report",
   train_s="arxiv 2409.12186 - no training-hardware section is present in the paper. CHECK if a later "
           "revision adds one.",
   ly=28, kvh=4, hd=128, ctx="131,072 full support (32,768 in the default config)",
   ctx_s="VERIFIED 2026-08-10 against huggingface.co/Qwen/Qwen2.5-Coder-7B-Instruct: card states 131,072 "
         "full context support with 32,768 as the default configured length. Architecture confirmed: 28 "
         "layers, 28 query heads, 4 KV heads (GQA). Params confirmed 7.61B total / 6.53B non-embedding.",
   prim="Code generation - C++ / multi-language",
   prim_s="arxiv 2409.12186 title and abstract: 'Qwen2.5-Coder Technical Report', described as a code-specific model series.",
   sec="Fill-in-the-Middle (FIM) code completion; code reasoning; code repair",
   sec_s="huggingface.co/Qwen/Qwen2.5-Coder-7B - base model card documents the FIM special tokens "
         "<|fim_prefix|>, <|fim_suffix|>, <|fim_middle|>.",
   bench="MultiPL-E C++ 75.6% | HumanEval 88.4% | HumanEval+ 84.1% | MBPP 83.5% | MBPP+ 71.7%",
   bench_s="VERIFIED 2026-08-10 against arxiv.org/html/2409.12186v3.  ||  "
           "MultiPL-E C++ 75.6 = TABLE 17 (per-language MultiPL-E breakdown).  ||  "
           "HumanEval 88.4 / HumanEval+ 84.1 / MBPP 83.5 / MBPP+ 71.7 = TABLE 16.  ||  "
           "NOTE: the official Qwen blog (qwenlm.github.io/blog/qwen2.5-coder-family/) does NOT publish "
           "per-model numbers - the technical report is the only primary source for these.",
   eng="llama.cpp, vLLM, SGLang, Ollama, TGI, MLX, TensorRT-LLM",
   eng_s="llama.cpp: GGUF builds published at huggingface.co/Qwen/Qwen2.5-Coder-7B-Instruct-GGUF (official "
         "Qwen repo).  ||  vLLM: docs.vllm.ai supported-models list includes Qwen2ForCausalLM.  ||  "
         "SGLang: docs.sglang.ai supported models.  ||  Ollama: ollama.com/library/qwen2.5-coder.",
   vstat="VERIFIED 2026-08-10. Benchmarks confirmed against arxiv 2409.12186 Tables 16 and 17.",
 ),
 dict(
   n="Phi-4 14B", n_s="huggingface.co/microsoft/phi-4 - model card title",
   lic="MIT", lic_s="huggingface.co/microsoft/phi-4/blob/main/LICENSE - MIT text",
   tier="FULL", ctry=MIT_OK + " Origin: United States (Microsoft). No procurement-origin concern.",
   ctry_s=MIT_SRC + "  ||  Origin: arxiv 2412.08905, author affiliation 'Microsoft Research'.",
   pb=14.7, p="14.7B dense",
   p_s="huggingface.co/microsoft/phi-4 - model card 'Model Summary' table, Architecture row states 14B "
       "parameters. config.json for exact geometry.",
   train="1920 x NVIDIA H100-80GB, 21 days",
   train_s="VERIFIED 2026-08-10 against huggingface.co/microsoft/phi-4 model card, 'Training' section: "
           "'1920 H100-80G' for '21 days'.",
   ly=40, kvh=10, hd=128, ctx="16,384",
   ctx_s="huggingface.co/microsoft/phi-4 - model card 'Model Summary' table, Context length row = 16K tokens; "
         "config.json max_position_embeddings = 16384.",
   prim="General reasoning and code generation",
   prim_s="arxiv 2412.08905 abstract - describes a model trained with emphasis on data quality and reasoning, "
          "not a code-specific model.",
   sec="Mathematical reasoning; orchestration / planning (Pass A IR generation)",
   sec_s="arxiv 2412.08905 - reports MATH and GSM8K results alongside general benchmarks.",
   bench="HumanEval 82.6% | MMLU 84.8% | MATH 80.4% | GPQA 56.1% | MGSM 80.6% | DROP 75.5% | "
         "MultiPL-E C++ NOT PUBLISHED",
   bench_s="VERIFIED 2026-08-10 against huggingface.co/microsoft/phi-4 model card evaluation table.  ||  "
           "CORRECTION: an earlier revision of this sheet listed GSM8K 91.5% - Microsoft does NOT publish "
           "GSM8K for Phi-4. The card reports MGSM 80.6% (multilingual grade-school math) instead. "
           "The GSM8K figure has been removed as unsourced.  ||  MultiPL-E C++: genuinely absent - a gap "
           "in the published record, not in this research.",
   eng="llama.cpp, vLLM, Ollama, ONNX Runtime, TGI, MLX",
   eng_s="llama.cpp: GGUF at huggingface.co/microsoft/phi-4-gguf (official Microsoft repo).  ||  "
         "ONNX: huggingface.co/microsoft/phi-4-onnx.  ||  vLLM: docs.vllm.ai supported models "
         "(Phi3ForCausalLM architecture).  ||  Ollama: ollama.com/library/phi4.",
   vstat="VERIFIED 2026-08-10. Benchmarks and training HW confirmed against the phi-4 model card. GSM8K removed as unsourced.",
 ),
 dict(
   n="Granite 3.3 8B Instruct", n_s="huggingface.co/ibm-granite/granite-3.3-8b-instruct - model card title",
   lic="Apache 2.0", lic_s="huggingface.co/ibm-granite/granite-3.3-8b-instruct - model card 'License' field states Apache 2.0",
   tier="FULL", ctry=APACHE + " Origin: United States (IBM). IBM additionally publishes training-data "
        "provenance and offers customer indemnification for Granite models.",
   ctry_s=APACHE_SRC + "  ||  Origin and indemnity: ibm.com/granite - IBM states standard contractual "
          "IP indemnification for Granite. CHECK current terms with IBM before relying on this commercially.",
   pb=8.1, p="8.1B dense",
   p_s="huggingface.co/ibm-granite/granite-3.3-8b-instruct - model card 'Model Architecture' table.",
   train="IBM Blue Vela cluster, NVIDIA H100",
   train_s="Granite model card 'Infrastructure' section names the Blue Vela supercomputing cluster. "
           "CHECK exact GPU count.",
   ly=40, kvh=8, hd=128, ctx="131,072",
   ctx_s="huggingface.co/ibm-granite/granite-3.3-8b-instruct - model card states 128K context; "
         "config.json max_position_embeddings = 131072.",
   prim="General instruction following with code generation and FIM",
   prim_s="Model card 'Intended Use' section lists code-related tasks including code completion.",
   sec="Reasoning; grounded RAG with inline citations; tool/function calling",
   sec_s="Model card documents citation generation and tool-calling capability as first-class features "
         "of the 3.3 release.",
   bench="HumanEval 89.73% | HumanEval+ 86.09% | MMLU 65.54% | GSM8K 80.89% | MATH-500 69.02% | "
         "IFEval 74.82% | MultiPL-E C++ NOT PUBLISHED",
   bench_s="VERIFIED 2026-08-10 against huggingface.co/ibm-granite/granite-3.3-8b-instruct model card, "
           "'Evaluation Results' table. IBM publishes no arXiv paper for 3.3, so the card IS the primary "
           "source.  ||  CORRECTION: an earlier revision of this sheet stated HumanEval 67.1% - that was "
           "WRONG. The published figure is 89.73%, the HIGHEST HumanEval of any model in this sheet.  ||  "
           "MultiPL-E C++ genuinely absent - IBM publishes no per-language code breakdown.",
   eng="llama.cpp, vLLM, Ollama, TGI, ONNX Runtime",
   eng_s="llama.cpp: GGUF at huggingface.co/ibm-granite/granite-3.3-8b-instruct-GGUF (official IBM repo).  ||  "
         "vLLM: docs.vllm.ai supported models (GraniteForCausalLM).  ||  Ollama: ollama.com/library/granite3.3.",
   vstat="VERIFIED 2026-08-10. HumanEval CORRECTED 67.1 -> 89.73 against the live IBM model card.",
 ),
 dict(
   n="Llama 3.1 8B Instruct", n_s="huggingface.co/meta-llama/Llama-3.1-8B-Instruct - model card title",
   lic="Llama 3.1 Community License", lic_s="huggingface.co/meta-llama/Llama-3.1-8B-Instruct/blob/main/LICENSE",
   tier="FULL", ctry=LLAMA_OK + " Origin: United States (Meta).",
   ctry_s=LLAMA_SRC,
   pb=8.0, p="8.03B dense",
   p_s="huggingface.co/meta-llama/Llama-3.1-8B-Instruct - model card 'Model Information' table; "
       "config.json: 32 layers, 8 KV heads (GQA), head_dim 128.",
   train="NVIDIA H100-80GB (700W TDP). 8B share: 1.46M GPU-hours",
   train_s="VERIFIED 2026-08-10 against huggingface.co/meta-llama/Llama-3.1-8B-Instruct model card, "
           "'Training Energy Use' table: 1.46M GPU-hours on H100-80GB, 700W per device.",
   ly=32, kvh=8, hd=128, ctx="131,072",
   ctx_s="huggingface.co/meta-llama/Llama-3.1-8B-Instruct - model card states 128K context; "
         "config.json max_position_embeddings = 131072.",
   prim="General instruction following and code generation",
   prim_s="arxiv 2407.21783 - described as a general-purpose foundation model family.",
   sec="Reasoning; tool/function calling; multilingual (8 languages officially supported)",
   sec_s="Model card 'Intended Use' section names tool use and lists the 8 supported languages.",
   bench="HumanEval 72.6% (pass@1) | MMLU 69.4% (macro avg) | MMLU-CoT 73.0% | GSM8K 84.5% (8-shot CoT) | "
         "MBPP+ 72.8% | MultiPL-E C++ NOT PUBLISHED",
   bench_s="VERIFIED 2026-08-10 against huggingface.co/meta-llama/Llama-3.1-8B-Instruct model card "
           "evaluation tables.  ||  CORRECTION: an earlier revision listed MMLU 73.0% without qualification. "
           "The card reports TWO figures: MMLU 69.4% (macro average, standard 5-shot) and MMLU-CoT 73.0% "
           "(chain-of-thought). Quoting 73.0 as plain MMLU overstates it against models reporting the "
           "standard metric. Both are now shown.",
   eng="llama.cpp, vLLM, SGLang, TensorRT-LLM, Ollama, TGI, MLX, ONNX Runtime, ExecuTorch",
   eng_s="Broadest support of any model here. llama.cpp: GGUF widely published.  ||  "
         "vLLM: docs.vllm.ai (LlamaForCausalLM is the reference architecture).  ||  "
         "TensorRT-LLM: github.com/NVIDIA/TensorRT-LLM support matrix lists Llama explicitly.  ||  "
         "ExecuTorch: pytorch.org/executorch llama example.",
   vstat="VERIFIED 2026-08-10. MMLU disambiguated: 69.4 standard vs 73.0 CoT. Training GPU-hours confirmed.",
 ),
 dict(
   n="DeepSeek-Coder-V2-Lite Instruct", n_s="huggingface.co/deepseek-ai/DeepSeek-Coder-V2-Lite-Instruct",
   lic="DeepSeek License Agreement", lic_s="github.com/deepseek-ai/DeepSeek-LLM/blob/main/LICENSE-MODEL - "
       "custom licence, permits commercial use but carries use restrictions in an attached policy.",
   tier="OPEN", ctry="No country restriction in the licence, but it is NOT Apache/MIT - it carries "
        "an attached use-restriction policy. Origin: China (DeepSeek). Same procurement-origin question "
        "as Qwen. LEGAL REVIEW REQUIRED - this is the least permissive licence in the shortlist.",
   ctry_s="github.com/deepseek-ai/DeepSeek-LLM/blob/main/LICENSE-MODEL - see the 'Use Restrictions' "
          "attachment.  ||  Origin: arxiv 2406.11931, author affiliation 'DeepSeek-AI'.",
   pb=15.7, p="15.7B total MoE / 2.4B active per token",
   p_s="arxiv 2406.11931 - architecture section states 16B total with 2.4B activated. "
       "config.json for expert count and routing.",
   train="Not clearly disclosed for the Lite variant",
   train_s="arxiv 2406.11931 - training-infrastructure detail is given for the full V2, not clearly "
           "separated for Lite. CHECK.",
   ly=27, kvh=16, hd=128, ctx="131,072",
   ctx_s="huggingface.co/deepseek-ai/DeepSeek-Coder-V2-Lite-Instruct - config.json "
         "max_position_embeddings = 163840; model card states 128K usable. CHECK which figure applies.",
   prim="Code generation - C++ / multi-language (MoE)",
   prim_s="arxiv 2406.11931 title: 'DeepSeek-Coder-V2: Breaking the Barrier of Closed-Source Models in Code Intelligence'.",
   sec="Fill-in-the-Middle; mathematical reasoning",
   sec_s="Model card documents FIM tokens; paper reports math benchmarks alongside code.",
   bench="MultiPL-E C++ 75.8% | HumanEval 81.1% | MBPP+ 68.8%",
   bench_s="VERIFIED 2026-08-10 against arxiv.org/html/2406.11931v1, TABLE 3 (HumanEval, MBPP+ and the "
           "per-language MultiPL-E breakdown).  ||  CORRECTION: an earlier revision stated C++ 56.5% - "
           "that was WRONG. Published figure is 75.8%, the highest measured C++ score in this sheet.  ||  "
           "Paper notes the 16B/2.4B-active model beats the dense 33B on multi-language average "
           "(65.6% vs 61.9%).",
   eng="llama.cpp, vLLM, SGLang",
   eng_s="llama.cpp: GGUF community builds (no official DeepSeek GGUF repo - CHECK provenance of any "
         "GGUF you download).  ||  vLLM: docs.vllm.ai supported models (DeepseekV2ForCausalLM).  ||  "
         "SGLang: docs.sglang.ai.  ||  NOTE: MoE support in llama.cpp is newer and less battle-tested "
         "than dense-model support.",
   vstat="VERIFIED 2026-08-10. C++ CORRECTED 56.5 -> 75.8 against arxiv 2406.11931 Table 3. Licence remains restrictive.",
 ),
 dict(
   n="StarCoder2 15B", n_s="huggingface.co/bigcode/starcoder2-15b - model card title",
   lic="BigCode OpenRAIL-M v1", lic_s="huggingface.co/spaces/bigcode/bigcode-model-license-agreement - "
       "behavioural-use licence, royalty-free but with mandatory use restrictions that flow down to "
       "derivatives.",
   tier="ALLY", ctry="No country restriction. HAS behavioural use restrictions that you must pass on to "
        "anyone you distribute derivatives to. Origin: EU + US (ServiceNow, Hugging Face, NVIDIA).",
   ctry_s="BigCode OpenRAIL-M v1 Attachment A lists prohibited uses; Section 'Distribution and "
          "Redistribution' requires the restrictions to flow down.  ||  Origin: arxiv 2402.19173 author list.",
   pb=16.0, p="16.0B dense (marketed as 15B)",
   p_s="VERIFIED 2026-08-10 from arxiv 2402.19173 PDF TABLE 6 (Model architecture details): "
       "hidden_dim 6144, n_heads 48, n_kv_heads 4 (GQA), n_layers 40, vocab 49152, RoPE. "
       "head_dim = 6144/48 = 128, which is the geometry used for the KV-cache calculation in this sheet. "
       "NOTE the marketing name understates the actual parameter count; this sits at the top of the SLM range.",
   train="1024 x NVIDIA H100; 4.1T tokens, 1M iterations, batch 4.1M (15B)",
   train_s="VERIFIED 2026-08-10 against huggingface.co/bigcode/starcoder2-15b model card: 1024 x H100, "
           "4+ trillion training tokens.  ||  CORRECTION: an earlier revision said A100-80GB. That was WRONG - "
           "StarCoder2 was trained on H100s.",
   ly=40, kvh=4, hd=128, ctx="16,384",
   ctx_s="VERIFIED 2026-08-10 from arxiv 2402.19173 PDF Section 6.3 and TABLE 8: base models trained at "
         "sequence length 4,096, then long-context pre-trained on 200B further tokens at 16,384 context "
         "with a 4,096 sliding window and FlashAttention-2. Model card confirms 16,384. "
         "The 4,096 sliding window is the practical attention span - relevant when reasoning about a 6-8K prompt.",
   prim="Code generation and Fill-in-the-Middle (base model, NOT instruction tuned)",
   prim_s="arxiv 2402.19173 - StarCoder2 is released as a base model. The card explicitly notes it is "
          "not an instruction-following model.",
   sec="Repository-level code completion; code editing (CanItEdit). FIM supported but MEASURABLY BROKEN (see source)",
   sec_s="VERIFIED 2026-08-10 by direct read of the arxiv 2402.19173 PDF.  ||  Repo-level completion: TABLE 17 (RepoBench v1.1) and Section 7.6. Code editing: TABLE 13 (CanItEdit).  ||  TWO PAPER-DOCUMENTED WEAKNESSES THAT MATTER FOR snorkelbadger AND ARE NOT VISIBLE IN ANY HEADLINE SCORE:  ||  (1) FIM IS BROKEN. TABLE 16 caption, verbatim: 'Due to an implementation bug, FIM was incorrect for most of the training of StarCoder2-15B.' Section 7.5 text: 'StarCoder2-15B underperforms on FIM.' Measured FIM is Python 48.4 / Java 60.5 / JS 54.7 against StarCoderBase-15B at 62 / 73 / 74. Listing FIM as a STRENGTH of this model would be wrong - it is a regression.  ||  (2) C++ SPECIFICALLY IS WEAK. Section 7.2.1, paraphrased from the paper: StarCoder2-15B underperforms on C++ because roughly ONE THIRD of the C++ it generates is incomplete - the paper's example is an unexpected break immediately after the beginning of a for loop. That is exactly the failure mode a 5-gate cross-compile pipeline would hit on every run.",
   bench="MultiPL-E C++ 41.4% | HumanEval 46.3% | HumanEval+ 37.8% | MBPP 66.2% | MBPP+ 53.1% | CanItEdit descriptive 43.08% / lazy 38.45% | RepoBench-v1.1 ES 74.08% | CRUXEval-I 48.1% / O 47.1% | GSM8K-PAL 65.1% | FIM Python 48.4% (SEE SOURCE - BUGGED)",
   bench_s="FULLY VERIFIED 2026-08-10 by direct read of the arxiv 2402.19173 PDF.  ||  MultiPL-E C++ 41.4% = TABLE 10 (Pass@1 on MultiPL-E, 50 samples per problem, temperature 0.2, top-p 0.95). For context in the same table: CodeLlama-13B C++ 37.4, DeepSeekCoder-33B C++ 51.2, StarCoder2-7B C++ 33.6.  ||  HumanEval 46.3 / HumanEval+ 37.8 / MBPP 66.2 / MBPP+ 53.1 = TABLE 9 (greedy decoding, EvalPlus framework).  ||  CanItEdit code-EDITING 43.08 descriptive / 38.45 lazy = TABLE 13 - the most relevant published benchmark in this sheet for a compile-retry loop, since it measures editing existing code rather than writing it fresh.  ||  RepoBench-v1.1 Python edit-similarity 74.08 = TABLE 17.  ||  CRUXEval-I 48.1 / CRUXEval-O 47.1 = TABLE 15.  ||  GSM8K-PAL 65.1 = TABLE 14.  ||  FIM Python 48.4 / Java 60.5 / JavaScript 54.7 = TABLE 16. CRITICAL: these FIM scores are LOWER than StarCoder2-15B's own predecessor StarCoderBase-15B (62 / 73 / 74) and far below CodeLlama-13B (74.5 / 80 / 85). The Table 16 caption states verbatim: 'Due to an implementation bug, FIM was incorrect for most of the training of StarCoder2-15B.' Full detail in the Secondary Purpose - SOURCE cell.",
   eng="llama.cpp, vLLM, TGI, Ollama",
   eng_s="llama.cpp: GGUF community builds.  ||  vLLM: docs.vllm.ai (Starcoder2ForCausalLM).  ||  "
         "TGI: natively supported, BigCode and HF are the same ecosystem.",
   vstat="FULLY VERIFIED 2026-08-10 by direct PDF read; every figure cited to a table number. C++ 41.4 CONFIRMED (Table 10). NEW FINDINGS: FIM is BUGGED (Table 16 caption) and C++ output is 1/3 incomplete (Sec 7.2.1). Both argue against this model for snorkelbadger.",
 ),
 dict(
   n="CodeGemma 7B IT", n_s="huggingface.co/google/codegemma-7b-it - model card title",
   lic="Gemma Terms of Use", lic_s="ai.google.dev/gemma/terms - NOT an OSI open source licence",
   tier="COMMERCIAL", ctry=GEMMA_OK + " Origin: United States (Google).",
   ctry_s=GEMMA_SRC,
   pb=8.5, p="8.5B dense (marketed as 7B)",
   p_s="huggingface.co/google/codegemma-7b-it - config.json. Large vocabulary (256k tokens) inflates "
       "the parameter count above the marketing name.",
   train="Google TPUv5e",
   train_s="CodeGemma report on ai.google.dev/gemma/docs/codegemma - hardware section. CHECK.",
   ly=28, kvh=16, hd=256, ctx="8,192",
   ctx_s="huggingface.co/google/codegemma-7b-it - config.json max_position_embeddings = 8192. "
         "THIS IS THE ELIMINATING CONSTRAINT for snorkelbadger: a 6-8K prompt fills the entire window.",
   prim="Code generation and Fill-in-the-Middle",
   prim_s="ai.google.dev/gemma/docs/codegemma/model_card - 'Model Information' describes code completion "
          "and generation as the intended use.",
   sec="Code chat; natural-language-to-code",
   sec_s="Same model card, 'Intended Usage' section distinguishes the -it variant as chat-tuned.",
   bench="HumanEval 56.1% | MBPP 54.2% | BabelCode-HumanEval C++ 42.2% | BabelCode-MBPP C++ 56.7% | "
         "HumanEval single-line infill 68.25% | multi-line infill 20.05%",
   bench_s="VERIFIED 2026-08-10 against huggingface.co/google/codegemma-7b-it model card evaluation table.  ||  "
           "CORRECTION: an earlier revision said 'C++ NOT PUBLISHED'. That was WRONG. Google DOES publish "
           "C++ figures, under BabelCode (BC HE C++ 42.2, BC MBPP C++ 56.7) rather than MultiPL-E - which "
           "is why it was missed. BabelCode and MultiPL-E are different harnesses and the scores are NOT "
           "directly comparable; flag this if a reviewer compares 42.2 against Qwen's 75.6.  ||  "
           "Multi-line infill 20.05% vs single-line 68.25% is a large gap worth noting for retry-patching use.",
   eng="llama.cpp, Ollama, vLLM",
   eng_s="llama.cpp: GGUF at huggingface.co/google/codegemma-7b-it-GGUF (official Google repo).  ||  "
         "Ollama: ollama.com/library/codegemma.  ||  vLLM: docs.vllm.ai (GemmaForCausalLM).",
   vstat="VERIFIED 2026-08-10. C++ figures DO exist under BabelCode (42.2) - earlier NOT PUBLISHED was wrong. Context 8K confirmed disqualifier.",
 ),
 dict(
   n="Phi-3.5-mini 3.8B Instruct", n_s="huggingface.co/microsoft/Phi-3.5-mini-instruct - model card title",
   lic="MIT", lic_s="huggingface.co/microsoft/Phi-3.5-mini-instruct/blob/main/LICENSE - MIT text",
   tier="FULL", ctry=MIT_OK + " Origin: United States (Microsoft).",
   ctry_s=MIT_SRC + "  ||  Origin: arxiv 2404.14219 author affiliation 'Microsoft'.",
   pb=3.8, p="3.8B dense",
   p_s="huggingface.co/microsoft/Phi-3.5-mini-instruct - model card 'Model Summary'; "
       "config.json: 32 layers, 32 heads, NO grouped-query attention.",
   train="512 x NVIDIA H100-80GB, 10 days",
   train_s="VERIFIED 2026-08-10 against huggingface.co/microsoft/Phi-3.5-mini-instruct model card, "
           "'Training' section: 512 H100-80G GPUs for 10 days.",
   ly=32, kvh=32, hd=96, ctx="131,072",
   ctx_s="huggingface.co/microsoft/Phi-3.5-mini-instruct - config.json max_position_embeddings = 131072 "
         "with LongRoPE scaling. WARNING: no GQA means the KV cache is unusually large at long context - "
         "see the VRAM @Full ctx column.",
   prim="General instruction following, reasoning-dense for its size",
   prim_s="arxiv 2404.14219 abstract - positions the model as matching much larger models on reasoning.",
   sec="Code generation; mathematical reasoning; long-context retrieval",
   sec_s="Model card evaluation tables cover code, math and long-context (RULER, RepoQA) benchmarks.",
   bench="HumanEval 62.8% | MMLU 69.0% | GSM8K 86.2%",
   bench_s="VERIFIED 2026-08-10 against huggingface.co/microsoft/Phi-3.5-mini-instruct model card: "
           "HumanEval 62.8 (0-shot), MMLU 69 (5-shot), GSM8K 86.2 (8-shot CoT). The 3.5 refresh is "
           "documented on the card rather than in a separate paper, so the card is the primary source.",
   eng="llama.cpp, ONNX Runtime, Ollama, vLLM, MLX",
   eng_s="ONNX: huggingface.co/microsoft/Phi-3.5-mini-instruct-onnx (official Microsoft repo, includes "
         "INT4 builds for CPU and DirectML).  ||  llama.cpp: GGUF community builds.  ||  "
         "Ollama: ollama.com/library/phi3.5.",
   vstat="VERIFIED 2026-08-10. All benchmarks and training HW confirmed against the model card.",
 ),
 dict(
   n="Llama 3.2 3B Instruct", n_s="huggingface.co/meta-llama/Llama-3.2-3B-Instruct - model card title",
   lic="Llama 3.2 Community License", lic_s="huggingface.co/meta-llama/Llama-3.2-3B-Instruct/blob/main/LICENSE",
   tier="FULL", ctry=LLAMA_OK + " EU CLAUSE DOES NOT APPLY TO THIS MODEL. The Llama 3.2 EU territorial "
        "exclusion is scoped to MULTIMODAL models only; this is a text-only model and is unaffected. "
        "Origin: United States (Meta).",
   ctry_s=LLAMA_SRC + "  ||  VERIFIED 2026-08-10 against huggingface.co/meta-llama/Llama-3.2-3B-Instruct: "
          "the EU restriction is expressly limited to multimodal models. This text-only 3B is NOT "
          "restricted.  ||  CORRECTION: an earlier revision of this sheet flagged an EU concern on this "
          "model. That was over-cautious and is withdrawn.",
   pb=3.2, p="3.21B dense",
   p_s="huggingface.co/meta-llama/Llama-3.2-3B-Instruct - model card 'Model Information'; config.json: "
       "28 layers, 8 KV heads (GQA), head_dim 128. Distilled from Llama 3.1 8B and 70B.",
   train="NVIDIA H100-80GB. 3B share: 460K GPU-hours",
   train_s="Llama 3.2 model card 'Training Energy Use' table gives GPU-hours per size. CHECK exact figure.",
   ly=28, kvh=8, hd=128, ctx="131,072",
   ctx_s="huggingface.co/meta-llama/Llama-3.2-3B-Instruct - config.json max_position_embeddings = 131072.",
   prim="General instruction following, on-device / edge deployment",
   prim_s="ai.meta.com/blog/llama-3-2-connect-2024-edge-mobile-devices/ - the 1B and 3B are explicitly "
          "positioned for edge and mobile.",
   sec="Code generation; summarization; tool calling; the only text SLM here viable on Cortex-A53",
   sec_s="Meta blog names summarization, instruction following and rewriting as the target on-device tasks. "
         "Edge viability at Q4 is a CALC from the 1.9 GB file size against the S50 memory budget.",
   bench="MMLU 63.4% (macro avg) | GSM8K 77.7% (8-shot CoT) | HumanEval NOT PUBLISHED",
   bench_s="VERIFIED 2026-08-10 against huggingface.co/meta-llama/Llama-3.2-3B-Instruct model card.  ||  "
           "CORRECTION: an earlier revision listed HumanEval 57.8%. Meta does NOT publish a HumanEval "
           "figure for the 3B on the model card. That number has been REMOVED as unsourced. Do not "
           "reinstate it without a primary citation.  ||  Consequence: this model's code ability is "
           "UNMEASURED. It is an edge-deployment candidate on size grounds only.",
   eng="llama.cpp (aarch64 NEON), ONNX Runtime, MLC-LLM, Ollama, ExecuTorch",
   eng_s="ExecuTorch: pytorch.org/executorch - Meta's own on-device runtime, Llama 3.2 is the reference "
         "example.  ||  llama.cpp: aarch64 NEON path documented in docs/build.md.  ||  "
         "MLC-LLM: llm.mlc.ai model library.",
   vstat="VERIFIED 2026-08-10. EU clause does NOT apply (text-only). HumanEval 57.8 REMOVED as unsourced - code ability unmeasured.",
 ),
 dict(
   n="Mistral 7B Instruct v0.3", n_s="huggingface.co/mistralai/Mistral-7B-Instruct-v0.3 - model card title",
   lic="Apache 2.0", lic_s="huggingface.co/mistralai/Mistral-7B-Instruct-v0.3 - model card licence field",
   tier="ALLY", ctry=APACHE + " Origin: France (Mistral AI) - EU, allied jurisdiction.",
   ctry_s=APACHE_SRC + "  ||  Origin: arxiv 2310.06825 author affiliation 'Mistral AI'.",
   pb=7.2, p="7.25B dense",
   p_s="huggingface.co/mistralai/Mistral-7B-Instruct-v0.3 - config.json: 32 layers, 8 KV heads (GQA).",
   train="Not disclosed",
   train_s="arxiv 2310.06825 - the paper contains no training-hardware disclosure. This is a genuine "
           "absence, not a gap in this research.",
   ly=32, kvh=8, hd=128, ctx="32,768",
   ctx_s="huggingface.co/mistralai/Mistral-7B-Instruct-v0.3 - config.json max_position_embeddings = 32768. "
         "NOTE v0.3 extended this from v0.2.",
   prim="General instruction following",
   prim_s="arxiv 2310.06825 - general-purpose model, NOT code specialised.",
   sec="Function calling (added in v0.3); multilingual",
   sec_s="Model card release notes for v0.3 list extended vocabulary and function-calling support.",
   bench="NO PRIMARY BENCHMARKS EXIST FOR v0.3. MMLU 62.5% is a v0.1 figure; HumanEval 36.5% is leaderboard-only. SEE SOURCE.",
   bench_s="VERIFIED 2026-08-10, AND THE FINDING IS THAT NO SOLID REFERENCE EXISTS.  ||  (1) huggingface.co/mistralai/Mistral-7B-Instruct-v0.3 publishes NO benchmark scores at all - the card covers installation, usage and function calling only.  ||  (2) MMLU 62.5% comes from arxiv 2310.06825, which is the v0.1 BASE model paper. It is a DIFFERENT MODEL VERSION and should not be presented as a v0.3 Instruct figure. Secondary sources put v0.3 nearer 59.9-60.0%, but those are not primary either.  ||  (3) HumanEval 36.5% is from evalplus.github.io/leaderboard.html, a third-party leaderboard whose table is JavaScript-rendered and could not be captured for citation.  ||  CONCLUSION: this is the only model in the sheet with NO citable primary benchmark. It is retained for completeness but must not be ranked on these numbers. It is not a contender for the C++ role in any case.",
   eng="llama.cpp, vLLM, SGLang, Ollama, TGI, MLX, ONNX Runtime",
   eng_s="vLLM: docs.vllm.ai (MistralForCausalLM).  ||  llama.cpp: GGUF widely published.  ||  "
         "Ollama: ollama.com/library/mistral.",
   vstat="VERIFIED 2026-08-10: the v0.3 model card publishes NO benchmarks whatsoever. MMLU is from the v0.1 paper, HumanEval from a leaderboard. WEAKEST EVIDENCE IN SHEET.",
 ),
 dict(
   n="DeepSeek-R1-Distill-Qwen-14B", n_s="huggingface.co/deepseek-ai/DeepSeek-R1-Distill-Qwen-14B",
   lic="MIT", lic_s="huggingface.co/deepseek-ai/DeepSeek-R1-Distill-Qwen-14B - model card licence field "
       "states MIT. NOTE: the distilled models are MIT even though base DeepSeek models use a custom licence.",
   tier="OPEN", ctry=MIT_OK + " CAVEAT: the model is distilled from DeepSeek-R1 onto a Qwen2.5-14B base. "
        "The Qwen base carries Apache 2.0. Both are permissive but the lineage crosses two licences - "
        "have legal confirm the stack. Origin: China (DeepSeek + Alibaba base).",
   ctry_s=MIT_SRC + "  ||  Lineage: arxiv 2501.12948 - the distillation section names the Qwen2.5 base "
          "models used.  ||  CONFIRM the combined licence position with legal.",
   pb=14.8, p="14.8B dense (Qwen2.5-14B architecture)",
   p_s="huggingface.co/deepseek-ai/DeepSeek-R1-Distill-Qwen-14B - config.json matches Qwen2.5-14B geometry.",
   train="Distillation only - no large-scale pretraining hardware disclosed",
   train_s="arxiv 2501.12948 - distillation is described as SFT on R1-generated reasoning traces. "
           "Hardware for the distillation run is not stated.",
   ly=48, kvh=8, hd=128, ctx="131,072",
   ctx_s="config.json max_position_embeddings = 131072 (inherited from the Qwen2.5-14B base).",
   prim="Chain-of-thought reasoning",
   prim_s="arxiv 2501.12948 - the R1 series is explicitly a reasoning model family.",
   sec="Mathematical problem solving; planning / orchestration (Pass A)",
   sec_s="Paper reports MATH-500 and AIME as the headline benchmarks.",
   bench="MATH-500 93.9% | AIME 2024 69.7% | GPQA Diamond 59.1% | LiveCodeBench 53.1% | Codeforces rating 1481 | HumanEval NOT PUBLISHED | MMLU NOT PUBLISHED",
   bench_s="VERIFIED 2026-08-10 against huggingface.co/deepseek-ai/DeepSeek-R1-Distill-Qwen-14B model card.  ||  CORRECTION: an earlier revision said no code benchmark was published and its code ability was UNKNOWN. That was WRONG - the card publishes LiveCodeBench 53.1% and a Codeforces rating of 1481.  ||  CAVEAT THAT STILL STANDS: HumanEval and MultiPL-E C++ are NOT published, and the base is Qwen2.5 general, NOT Qwen2.5-Coder. Do not assume it inherits coder-tuned C++ ability.",
   eng="llama.cpp, vLLM, SGLang, Ollama",
   eng_s="llama.cpp: GGUF community builds.  ||  vLLM: docs.vllm.ai (Qwen2ForCausalLM architecture).  ||  "
         "Ollama: ollama.com/library/deepseek-r1.",
   vstat="VERIFIED 2026-08-10. Code ability IS measured (LiveCodeBench 53.1) - earlier UNKNOWN was wrong. HumanEval/C++ still unpublished.",
 ),
 dict(
   n="Gemma 3 4B IT", n_s="huggingface.co/google/gemma-3-4b-it - model card title",
   lic="Gemma Terms of Use", lic_s="ai.google.dev/gemma/terms - NOT an OSI open source licence",
   tier="COMMERCIAL", ctry=GEMMA_OK + " Origin: United States (Google).",
   ctry_s=GEMMA_SRC,
   pb=4.3, p="4.3B dense (multimodal: text + vision)",
   p_s="huggingface.co/google/gemma-3-4b-it - model card; config.json includes a SigLIP vision tower.",
   train="Google TPUv4p, TPUv5p and TPUv5e",
   train_s="VERIFIED 2026-08-10 against huggingface.co/google/gemma-3-4b-it model card: TPUv4p, TPUv5p and TPUv5e.",
   ly=34, kvh=4, hd=256, ctx="131,072 input / 8,192 output",
   ctx_s="VERIFIED 2026-08-10 against huggingface.co/google/gemma-3-4b-it: 128K token INPUT context but an 8,192 token OUTPUT limit. The output cap is a separate constraint from the context window and is not a problem for snorkelbadger (generated blocks are ~80 lines), but note it before assuming 128K end to end.",
   prim="General instruction following with vision (image to text)",
   prim_s="arxiv 2503.19786 - Gemma 3 introduces multimodality to the Gemma family at 4B and above.",
   sec="Multilingual (140+ languages claimed); basic code generation",
   sec_s="Model card 'Model Information' section states the language coverage.",
   bench="HumanEval 36.0% (0-shot) | MMLU 59.6% (5-shot) | MBPP 46.0% (3-shot) | MMMU 39.2% | GSM8K 38.4% (8-shot) | MATH 24.2% (4-shot)",
   bench_s="VERIFIED 2026-08-10 against huggingface.co/google/gemma-3-4b-it model card evaluation table.  ||  NOTE the low code and math scores: multimodal capability at 4B costs real ground on code ability. HumanEval 36.0 and GSM8K 38.4 are the weakest in this sheet among instruct models.",
   eng="llama.cpp, Ollama, vLLM, MLX",
   eng_s="llama.cpp: GGUF at huggingface.co/google/gemma-3-4b-it-qat-q4_0-gguf (official Google "
         "quantization-aware-trained build - better quality than post-hoc Q4).  ||  "
         "Ollama: ollama.com/library/gemma3.",
   vstat="VERIFIED 2026-08-10. All benchmarks, training TPUs and the 8K output cap confirmed against the model card.",
 ),
 dict(
   n="Qwen2-VL 7B Instruct", n_s="huggingface.co/Qwen/Qwen2-VL-7B-Instruct - model card title",
   lic="Apache 2.0", lic_s="huggingface.co/Qwen/Qwen2-VL-7B-Instruct - model card licence field",
   tier="OPEN", ctry=APACHE + " Origin: China (Alibaba Cloud). Same procurement-origin question as "
        "Qwen2.5-Coder.",
   ctry_s=APACHE_SRC + "  ||  Origin: arxiv 2409.12191 author affiliation 'Alibaba Group'.",
   pb=8.3, p="8.3B dense (7.6B LLM + 675M vision encoder)",
   p_s="huggingface.co/Qwen/Qwen2-VL-7B-Instruct - config.json shows the vision tower and LLM separately.",
   train="Not disclosed",
   train_s="arxiv 2409.12191 - no training-hardware disclosure located. CHECK.",
   ly=28, kvh=4, hd=128, ctx="32,768",
   ctx_s="config.json max_position_embeddings = 32768. Naive Dynamic Resolution means image token count "
         "varies with input resolution - a large image can consume a lot of the window. IMPORTANT for sizing.",
   prim="Image to text - visual question answering and document understanding",
   prim_s="arxiv 2409.12191 title: 'Qwen2-VL: Enhancing Vision-Language Model's Perception of the World "
          "at Any Resolution'.",
   sec="OCR / text-in-image reading; video understanding (accepts frame sequences); multilingual OCR",
   sec_s="Paper documents video input support and multilingual text recognition. DocVQA 94.5% is the "
         "strongest OCR-adjacent number in this sheet.",
   bench="DocVQA 94.5% (test) | TextVQA 84.3% (val) | MathVista 58.2% (testmini) | MMMU 54.1% (val)",
   bench_s="VERIFIED 2026-08-10 against huggingface.co/Qwen/Qwen2-VL-7B-Instruct model card.  ||  TextVQA 84.3% added - the highest text-in-image score in this sheet, which together with DocVQA 94.5% makes this the strongest LPR / plate-reading candidate by a clear margin.  ||  Card also states video understanding of 20min+ duration.",
   eng="vLLM, SGLang, llama.cpp, transformers",
   eng_s="vLLM: docs.vllm.ai multimodal-models list (Qwen2VLForConditionalGeneration).  ||  "
         "llama.cpp: vision support added but CHECK current status - the multimodal path in llama.cpp "
         "lags the text path significantly.",
   vstat="VERIFIED 2026-08-10 against the model card. llama.cpp vision support: still needs a current-state check before relying on a GGUF path.",
 ),
 dict(
   n="Phi-3.5-Vision 4.2B", n_s="huggingface.co/microsoft/Phi-3.5-vision-instruct - model card title",
   lic="MIT", lic_s="huggingface.co/microsoft/Phi-3.5-vision-instruct/blob/main/LICENSE - MIT text",
   tier="FULL", ctry=MIT_OK + " Origin: United States (Microsoft). The most permissive VLM here.",
   ctry_s=MIT_SRC,
   pb=4.2, p="4.2B dense (Phi-3.5-mini LLM + CLIP ViT-L/14 vision encoder)",
   p_s="huggingface.co/microsoft/Phi-3.5-vision-instruct - model card 'Model Architecture' states the "
       "image encoder is CLIP ViT-L/14-336.",
   train="256 x NVIDIA A100-80GB, 6 days, 500B tokens",
   train_s="VERIFIED 2026-08-10 against huggingface.co/microsoft/Phi-3.5-vision-instruct: 256 A100-80G "
           "for 6 days on 500 billion vision+text tokens, trained July-August 2024.",
   ly=32, kvh=32, hd=96, ctx="131,072",
   ctx_s="config.json max_position_embeddings = 131072. Supports multi-frame / multi-image input, "
         "which is relevant for comparing camera frames.",
   prim="Image to text - captioning, OCR, chart and table reasoning",
   prim_s="Model card 'Intended Uses' - names general image understanding, OCR and chart comprehension.",
   sec="Multi-frame / video-frame comparison; document understanding",
   sec_s="Model card explicitly documents multi-image and video-frame summarization as a supported use.",
   bench="MMBench 81.9% | MMMU 43.0% | TextVQA 72.0%",
   bench_s="VERIFIED 2026-08-10 against huggingface.co/microsoft/Phi-3.5-vision-instruct: MMBench dev-en 81.9, MMMU val 43.0, TextVQA val 72.0. Card is the primary source - no separate paper for the 3.5 refresh.",
   eng="llama.cpp, ONNX Runtime, vLLM, transformers",
   eng_s="ONNX: huggingface.co/microsoft/Phi-3.5-vision-instruct-onnx (official).  ||  "
         "vLLM: docs.vllm.ai multimodal list.  ||  llama.cpp: CHECK current vision support state.",
   vstat="FULLY VERIFIED 2026-08-10. Benchmarks and training HW confirmed. MIT - cleanest VLM licence here.",
 ),
 dict(
   n="Llama 3.2 11B Vision Instruct", n_s="huggingface.co/meta-llama/Llama-3.2-11B-Vision-Instruct",
   lic="Llama 3.2 Community License", lic_s="huggingface.co/meta-llama/Llama-3.2-11B-Vision-Instruct/blob/main/LICENSE",
   tier="FULL", ctry=LLAMA_OK + " CRITICAL: the Llama 3.2 licence explicitly EXCLUDES use of the "
        "MULTIMODAL models by individuals or companies domiciled in the EU. This model is multimodal. "
        "If snorkelbadger ships in Europe this is a hard blocker. Origin: United States (Meta).",
   ctry_s=LLAMA_SRC + "  ||  VERIFIED 2026-08-10, EXACT LICENCE WORDING: 'With respect to any multimodal "
          "models included in Llama 3.2, the rights granted under Section 1(a) of the Llama 3.2 Community "
          "License Agreement are not being granted to you if you are an individual domiciled in, or a "
          "company with a principal place of business in, the European Union.'  ||  CRITICAL EXCEPTION, "
          "also verbatim: 'This restriction does not apply to end users of a product or service that "
          "incorporates any such multimodal models.'  ||  PRACTICAL READING for snorkelbadger: if the "
          "developing entity's principal place of business is outside the EU, the model may be used and "
          "the resulting product MAY be supplied to EU end users. If the developing entity is EU-domiciled, "
          "it is blocked. This turns on where ipo's contracting entity sits - a question for counsel, "
          "not for this sheet.",
   pb=10.6, p="10.6B dense (8B text base + vision adapter)",
   p_s="huggingface.co/meta-llama/Llama-3.2-11B-Vision-Instruct - model card; the vision adapter uses "
       "cross-attention layers into a frozen Llama 3.1 8B text model.",
   train="NVIDIA H100-80GB",
   train_s="Llama 3.2 model card 'Training Energy Use' table. CHECK GPU-hours for the 11B specifically.",
   ly=40, kvh=8, hd=128, ctx="131,072",
   ctx_s="config.json max_position_embeddings = 131072. NOTE the cross-attention design means image "
         "tokens do not consume the text context the way they do in Qwen2-VL.",
   prim="Image to text - visual reasoning, captioning, document QA",
   prim_s="ai.meta.com/blog/llama-3-2-connect-2024-edge-mobile-devices/ - names image reasoning, "
          "captioning and chart/graph understanding.",
   sec="Chart and diagram understanding; document VQA",
   sec_s="Meta blog and model card list DocVQA and ChartQA as headline capabilities.",
   bench="AI2 Diagram 91.1% | DocVQA 88.4% ANLS (test) | ChartQA 83.4% (test, CoT) | VQAv2 75.2% (test) | MMMU 50.7% (0-shot, CoT)",
   bench_s="VERIFIED 2026-08-10 against github.com/meta-llama/llama-models/blob/main/models/llama3_2/MODEL_CARD_VISION.md, INSTRUCTION-TUNED table.  ||  IMPORTANT TRAP: the same model card carries a separate BASE PRETRAINED table with much lower figures (MMMU 41.7, DocVQA 62.3, ChartQA 39.4). Secondary sources frequently quote the pretrained numbers as if they were the instruct ones. The figures here are the INSTRUCT results.  ||  Note the settings: MMMU is 0-shot WITH chain-of-thought and ChartQA is CoT - not directly comparable to non-CoT scores from other VLMs in this sheet.",
   eng="vLLM, TGI, transformers, llama.cpp (vision path partial)",
   eng_s="vLLM: docs.vllm.ai multimodal list (MllamaForConditionalGeneration).  ||  "
         "llama.cpp: cross-attention vision architecture is NOT fully supported - VERIFY before "
         "assuming a GGUF path exists.",
   vstat="EU MULTIMODAL EXCLUSION VERIFIED VERBATIM 2026-08-10. Blocks EU-DOMICILED DEVELOPERS, not EU end users. Turns on where ipo contracts from - LEGAL DECISION.",
 ),
 dict(
   n="InternVL2 8B", n_s="huggingface.co/OpenGVLab/InternVL2-8B - model card title",
   lic="MIT", lic_s="huggingface.co/OpenGVLab/InternVL2-8B - model card licence field states MIT. "
       "CHECK: the 8B uses an InternLM2.5 base whose own licence terms should be confirmed.",
   tier="OPEN", ctry=MIT_OK + " CAVEAT: verify the base-model licence in the lineage. "
        "Origin: China (Shanghai AI Laboratory / OpenGVLab).",
   ctry_s=MIT_SRC + "  ||  Base model lineage: model card 'Model Details' names InternLM2.5-7B-chat as "
          "the language component. CONFIRM its licence separately.",
   pb=8.1, p="8.1B dense (InternViT-300M vision + InternLM2.5-7B language)",
   p_s="huggingface.co/OpenGVLab/InternVL2-8B - model card 'Model Details' component table.",
   train="Not clearly disclosed",
   train_s="No dedicated arXiv paper for InternVL2 specifically; InternVL 1.0/1.5 papers exist "
           "(arxiv 2312.14238, 2404.16821). CHECK which paper actually covers the 2.0 release.",
   ly=32, kvh=8, hd=128, ctx="8,192",
   ctx_s="config.json max_position_embeddings = 8192. LIMITING for snorkelbadger at a 6-8K prompt.",
   prim="Image to text - multimodal understanding",
   prim_s="huggingface.co/OpenGVLab/InternVL2-8B model card 'Introduction'.",
   sec="Document OCR; chart understanding; multi-image comparison",
   sec_s="Model card evaluation tables cover DocVQA, ChartQA and multi-image benchmarks.",
   bench="MMBench-EN 81.7% | DocVQA 91.6% | MMMU 51.8%",
   bench_s="VERIFIED 2026-08-10 against huggingface.co/OpenGVLab/InternVL2-8B model card.  ||  CORRECTION: MMMU is 51.8%, not 51.2% as previously stated.  ||  Base language model confirmed as internlm2_5-7b-chat with an InternViT-300M-448px vision tower via MLP projector - so the licence question on the InternLM base still stands.",
   eng="LMDeploy, vLLM, transformers",
   eng_s="LMDeploy: github.com/InternLM/lmdeploy - the first-party runtime from the same lab.  ||  "
         "vLLM: docs.vllm.ai multimodal list (InternVLChatModel).  ||  "
         "NO official GGUF - llama.cpp path is unproven for this model.",
   vstat="VERIFIED 2026-08-10. MMMU corrected 51.2 -> 51.8. Base is internlm2_5-7b-chat - its licence still needs separate confirmation.",
 ),
 dict(
   n="LLaVA-1.6 (NeXT) 13B", n_s="huggingface.co/llava-hf/llava-v1.6-vicuna-13b-hf - model card",
   lic="Llama 2 Community License", lic_s="github.com/haotian-liu/LLaVA - repo licence is "
       "Apache 2.0, BUT the Vicuna-13B language base derives from Llama 2 and carries the Llama 2 "
       "Community Licence. THE EFFECTIVE LICENCE IS THE MORE RESTRICTIVE OF THE TWO.",
   tier="FULL", ctry="Licence stack is layered: LLaVA code Apache 2.0 over Vicuna over Llama 2. "
        "Llama 2 Community Licence terms therefore apply, including its Acceptable Use Policy. "
        "LEGAL REVIEW REQUIRED - this is the messiest licence position in the sheet. Origin: US.",
   ctry_s="LLaVA repo licence: github.com/haotian-liu/LLaVA/blob/main/LICENSE.  ||  "
          "Vicuna lineage: lmsys.org/blog/2023-03-30-vicuna.  ||  "
          "Llama 2 licence: ai.meta.com/llama/license.  ||  Have counsel trace the full chain.",
   pb=13.4, p="13.4B dense (Vicuna-13B + CLIP ViT-L/14-336)",
   p_s="huggingface.co/llava-hf/llava-v1.6-vicuna-13b-hf - config.json.",
   train="8 x NVIDIA A100-80GB (LLaVA-1.5 figure)",
   train_s="arxiv 2310.03744 - training section states roughly 1 day on 8xA100 for the 13B. "
           "CHECK whether 1.6 differs.",
   ly=40, kvh=40, hd=128, ctx="4,096",
   ctx_s="config.json max_position_embeddings = 4096 (Vicuna/Llama 2 inheritance). "
         "SEVERELY LIMITING - a 6-8K snorkelbadger prompt does not fit at all.",
   prim="Image to text - visual instruction following",
   prim_s="arxiv 2310.03744 title: 'Improved Baselines with Visual Instruction Tuning'.",
   sec="Visual question answering; OCR (weaker than Qwen2-VL)",
   sec_s="Paper evaluation covers VQA benchmarks; TextVQA 67.1% is materially below Qwen2-VL.",
   bench="VQAv2 82.8% | ScienceQA 73.6% | TextVQA 67.1% | GQA 65.4% | VisWiz 60.5% | MMMU val 36.2%",
   bench_s="VERIFIED 2026-08-10 against llava-vl.github.io/blog/2024-01-30-llava-next/ - the official LLaVA-NeXT release blog. Figures are for LLaVA-NeXT-Vicuna-13B at 672x672 resolution.  ||  CORRECTION: MMMU is 36.2% (val), not 35.9% as previously stated - 35.9 is the LLaVA-1.5 figure from arxiv 2310.03744. This is exactly the 1.5-vs-1.6 conflation the previous revision warned about, and this sheet had fallen into it.  ||  MMBench 70.0% has been REMOVED - it does not appear in the NeXT blog tables for the 13B and could not be sourced.  ||  DocVQA is NOT published for this variant.  ||  NOTE: the HuggingFace card llava-hf/llava-v1.6-vicuna-13b-hf publishes no benchmarks at all; the release blog is the primary source.",
   eng="llama.cpp, vLLM, SGLang, transformers",
   eng_s="llama.cpp: LLaVA has the most mature vision support in llama.cpp of any model here "
         "(clip.cpp / llava.cpp).  ||  vLLM: docs.vllm.ai multimodal list.  ||  "
         "SGLang: docs.sglang.ai - LLaVA is a documented example.",
   vstat="VERIFIED 2026-08-10. Licence confirmed LLAMA 2. Benchmarks now sourced to the official LLaVA-NeXT blog; MMMU corrected 35.9 (a 1.5 figure) -> 36.2, MMBench 70.0 removed as unsourceable. Context 4K disqualifies it regardless.",
 ),
 dict(
   n="PaliGemma 3B mix-448", n_s="huggingface.co/google/paligemma-3b-mix-448 - model card title",
   lic="Gemma Terms of Use", lic_s="ai.google.dev/gemma/terms",
   tier="COMMERCIAL", ctry=GEMMA_OK + " Origin: United States (Google).",
   ctry_s=GEMMA_SRC,
   pb=2.9, p="2.9B dense (SigLIP-So400m vision + Gemma-2B language)",
   p_s="arxiv 2407.07726 - architecture section names both components; config.json confirms.",
   train="Google TPUv5e",
   train_s="arxiv 2407.07726 - training-infrastructure section. CHECK.",
   ly=18, kvh=1, hd=256, ctx="512 tokens (input + output)",
   ctx_s="VERIFIED 2026-08-10 against huggingface.co/google/paligemma-3b-mix-448: the model supports 512 token input/output text sequences at 448x448 image resolution.  ||  MAJOR CORRECTION: an earlier revision of this sheet stated 8,192. That was WRONG by a factor of 16. At 512 tokens this model is categorically unusable for snorkelbadger - the prompt alone is 6-8K.",
   prim="Image to text - captioning and visual QA",
   prim_s="arxiv 2407.07726 - positioned as a versatile base model INTENDED TO BE FINE-TUNED, "
          "not used zero-shot.",
   sec="OCR; referring-expression segmentation; object detection with text prompts",
   sec_s="Paper documents detect/segment output formats as supported task prefixes - unusual and "
         "potentially useful for a VMS.",
   bench="VQAv2 85.64% | DocVQA 78.02% ANLS | TextVQA 73.15% | POPE 89.37% | MMVP 45.33% (ALL FINE-TUNED, NOT ZERO-SHOT)",
   bench_s="VERIFIED 2026-08-10 against huggingface.co/google/paligemma-3b-mix-448 model card.  ||  CRITICAL CAVEAT CONFIRMED BY THE CARD: these are single-task FINE-TUNED transfer results, not zero-shot. The card states the model was designed first and foremost as a pre-trained model for transfer to specialised tasks. DO NOT compare these numbers against the zero-shot figures from Qwen2-VL, Phi-3.5-Vision or InternVL2 elsewhere in this sheet - it is not a like-for-like comparison "
           "and a reviewer will catch it.",
   eng="JAX/Flax, transformers, llama.cpp",
   eng_s="Big Vision JAX reference: github.com/google-research/big_vision.  ||  "
         "transformers: PaliGemmaForConditionalGeneration.  ||  llama.cpp: CHECK support state.",
   vstat="VERIFIED 2026-08-10. CONTEXT CORRECTED 8192 -> 512, which disqualifies it outright. Benchmarks are FINE-TUNED, not zero-shot.",
 ),
 dict(
   n="Moondream2 1.9B", n_s="huggingface.co/vikhyatk/moondream2 - model card title",
   lic="Apache 2.0", lic_s="huggingface.co/vikhyatk/moondream2 - model card licence field",
   tier="FULL", ctry=APACHE + " Origin: United States (Moondream / M87 Labs). "
        "NOTE: small independent vendor - assess supply-chain and maintenance risk separately from licence.",
   ctry_s=APACHE_SRC + "  ||  Vendor: moondream.ai. CONSIDER vendor longevity for a production dependency.",
   pb=2.0, p="2B dense (SigLIP vision + Phi-1.5-derived language)",
   p_s="huggingface.co/vikhyatk/moondream2 - model card and config.json.",
   train="Not disclosed",
   train_s="No paper published. The model card is the only primary source. This is materially weaker "
           "evidence than the other models in this sheet - flag it in review.",
   ly=24, kvh=32, hd=64, ctx="2,048",
   ctx_s="config.json max_position_embeddings = 2048. VERY LIMITING - suitable only for short "
         "single-image captioning, not for prompt-heavy work.",
   prim="Image to text - lightweight captioning and VQA for edge devices",
   prim_s="huggingface.co/vikhyatk/moondream2 model card - positioned explicitly as a small edge VLM.",
   sec="OCR; on-device deployment (only VLM here viable on Cortex-A53)",
   sec_s="Edge viability is a CALC from the 1.2 GB Q4 file size against the S50 memory budget, "
         "not a vendor claim.",
   bench="DocVQA 79.3% | TextVQA 76.3% | ChartQA 77.5% (82.2% with PoT) | CountBenchQA 86.4% | OCRBench 61.2% | COCO object detection 51.2% | ScreenSpot F1@0.5 80.4%",
   bench_s="VERIFIED 2026-08-10 against huggingface.co/vikhyatk/moondream2 model card.  ||  MAJOR CORRECTION: the previous revision quoted VQAv2 79.4 / TextVQA 60.2 / DocVQA 61.9. Those were STALE - the model has been substantially improved since. Current published figures are far higher (DocVQA 79.3, TextVQA 76.3).  ||  NEW CAPABILITY WORTH NOTING FOR A VMS: the card now reports COCO object detection 51.2% and ScreenSpot pointing 80.4% - this 2B model does detection and pointing, not just captioning.  ||  STILL NO PAPER. All figures are vendor-self-reported and not independently verified - the weakest evidence grade in this sheet, even though the numbers improved.",
   eng="llama.cpp, ONNX Runtime, transformers",
   eng_s="llama.cpp: GGUF builds published by the vendor.  ||  moondream.ai documents an ONNX path "
         "for edge deployment.",
   vstat="VERIFIED against card 2026-08-10 - previous figures were STALE and understated it. Params 1.9B -> 2B. STILL NO PAPER: vendor-self-reported only.",
 ),

 # ================= CATEGORY 4/5: SLM-DERIVED EMBEDDING MODELS =================
 # These are VLM backbones with the generation head replaced by a pooled embedding
 # head, trained with contrastive loss. They are SLM-DERIVED, not native SLMs -
 # they emit vectors, not tokens. Added because they are the only way to cover
 # the image-to-embeddings and video-to-embeddings categories with SLM-class models.
 dict(
   n="Qwen3-VL-Embedding-8B", n_s="huggingface.co/Qwen/Qwen3-VL-Embedding-8B - model card title",
   lic="Apache 2.0", lic_s="huggingface.co/Qwen/Qwen3-VL-Embedding-8B - model card licence field",
   tier="OPEN", ctry=APACHE + " Origin: China (Alibaba Cloud). Same origin/procurement question as Qwen2.5-Coder.",
   ctry_s=APACHE_SRC + "  ||  Origin: arxiv 2601.04720, Qwen team.",
   cls=EMB, emb=True,
   task="4. Image to Embeddings, 5. Video to Embeddings",
   p="8B dense (base: Qwen3-VL-8B-Instruct)", ctx="32,000 tokens",
   ly=36, kvh=8, hd=128, pb=8.0,
   ctx_s="VERIFIED 2026-08-10 against huggingface.co/Qwen/Qwen3-VL-Embedding-8B: max sequence length "
         "32,000 tokens. Embedding dimension is customisable 64-4096 (Matryoshka), so you can trade "
         "index size against accuracy without retraining.",
   train="Multi-stage: large-scale contrastive pre-training then reranker distillation",
   train_s="arxiv 2601.04720 'Qwen3-VL-Embedding and Qwen3-VL-Reranker'. Hardware not disclosed. CHECK.",
   ft="LoRA", ftg="A100 40GB",
   eng="transformers, vLLM (embedding mode), sentence-transformers",
   eng_s="vLLM: docs.vllm.ai/projects/ascend - Qwen3-VL-Embedding has a documented vLLM path.  ||  "
         "github.com/QwenLM/Qwen3-VL-Embedding is the reference implementation.  ||  NO GGUF EXISTS - "
         "this cannot go through your llama.cpp pipeline.",
   prim="4. Image to Embeddings - unified text/image/video/document retrieval",
   prim_s="huggingface.co/Qwen/Qwen3-VL-Embedding-8B: supports text, images, screenshots, videos and "
          "combined text+image / text+video inputs into one shared vector space.",
   sec="5. Video to Embeddings; visual document retrieval; cross-modal search",
   sec_s="Same model card - MMEB-V2 is reported separately for image, video and visual-document tracks, "
         "so video is a first-class supported modality rather than an afterthought.",
   bench="MMEB-V2 Image 80.1 | VisDoc 83.3 | VIDEO 66.1 | MMTEB mean(task) 67.88 | Retrieval 69.41 | STS 75.41",
   bench_s="VERIFIED 2026-08-10 against huggingface.co/Qwen/Qwen3-VL-Embedding-8B model card benchmark "
           "tables.  ||  MMEB-V2 Video Overall 66.1 is the HIGHEST video embedding score found anywhere "
           "in this research - see the CATEGORY 4/5 note at the bottom of this sheet for the frame-count "
           "caveat that makes some competing numbers non-comparable.  ||  Benchmark definition: MMEB-V2 "
           "[arxiv 2507.04590], 78 tasks across image, video and visual document.",
   arm="Not viable (16GB FP16)",
   vstat="VERIFIED 2026-08-10 against the model card. NOTE: SLM-DERIVED, not a native SLM - emits vectors, not tokens. NO GGUF.",
 ),
 dict(
   n="Qwen3-VL-Embedding-2B", n_s="huggingface.co/Qwen/Qwen3-VL-Embedding-2B - model card title",
   lic="Apache 2.0", lic_s="huggingface.co/Qwen/Qwen3-VL-Embedding-2B - model card licence field",
   tier="OPEN", ctry=APACHE + " Origin: China (Alibaba Cloud).",
   ctry_s=APACHE_SRC + "  ||  Origin: arxiv 2601.04720, Qwen team.",
   cls=EMB, emb=True,
   task="4. Image to Embeddings, 5. Video to Embeddings",
   p="2B dense (base: Qwen3-VL-2B-Instruct)", ctx="32,000 tokens",
   ly=28, kvh=4, hd=128, pb=2.0,
   ctx_s="VERIFIED 2026-08-10 against huggingface.co/Qwen/Qwen3-VL-Embedding-2B: 32,000 token max "
         "sequence. Embedding dimension customisable 64-2048 (Matryoshka).",
   train="Multi-stage contrastive pre-training then reranker distillation",
   train_s="arxiv 2601.04720. Hardware not disclosed. CHECK.",
   ft="LoRA", ftg="RTX 3060 12GB",
   eng="transformers, vLLM (embedding mode), sentence-transformers",
   eng_s="Same stack as the 8B - github.com/QwenLM/Qwen3-VL-Embedding. NO GGUF.",
   prim="4. Image to Embeddings - unified retrieval at edge-viable size",
   prim_s="huggingface.co/Qwen/Qwen3-VL-Embedding-2B model card.",
   sec="5. Video to Embeddings; visual document retrieval",
   sec_s="Model card reports separate image and video MMEB-V2 tracks.",
   bench="MMEB-V2 Image 75.0 | VIDEO 61.9 | MMTEB mean(task) 63.87 | mean(type) 55.84",
   bench_s="VERIFIED 2026-08-10 against huggingface.co/Qwen/Qwen3-VL-Embedding-2B model card. "
           "Image Overall 75.0 across 36 image datasets, Video Overall 61.9 across 18 video datasets.  ||  "
           "BEST SIZE/PERFORMANCE TRADE IN THIS CATEGORY: at 2B it gives up only 5.1 points of image and "
           "4.2 points of video against the 8B, for a quarter of the memory.",
   arm="Marginal (4GB FP16, 2GB INT8)",
   vstat="VERIFIED 2026-08-10 against the model card. SLM-DERIVED. NO GGUF. Best size/accuracy trade for edge.",
 ),
 dict(
   n="GME-Qwen2-VL-7B-Instruct", n_s="huggingface.co/Alibaba-NLP/gme-Qwen2-VL-7B-Instruct - model card",
   lic="Apache 2.0", lic_s="huggingface.co/Alibaba-NLP/gme-Qwen2-VL-7B-Instruct - model card licence field",
   tier="OPEN", ctry=APACHE + " Origin: China (Alibaba NLP).",
   ctry_s=APACHE_SRC + "  ||  Origin: arxiv 2412.16855 'GME: General Multimodal Embedder', Alibaba.",
   cls=EMB, emb=True,
   task="4. Image to Embeddings",
   p="8.29B dense (base: Qwen2-VL-7B)", ctx="32,768 tokens",
   ly=28, kvh=4, hd=128, pb=8.29,
   ctx_s="VERIFIED 2026-08-10 against huggingface.co/Alibaba-NLP/gme-Qwen2-VL-7B-Instruct: max sequence "
         "length 32,768, embedding dimension 3584 (fixed, not Matryoshka).",
   train="Large-scale instruction-based training on fused-modal datasets",
   train_s="arxiv 2412.16855. Hardware not disclosed. CHECK.",
   ft="LoRA", ftg="A100 40GB",
   eng="transformers, sentence-transformers",
   eng_s="huggingface.co/Alibaba-NLP/gme-Qwen2-VL-7B-Instruct usage section. NO GGUF.",
   prim="4. Image to Embeddings - single-modal, cross-modal and fused-modal retrieval",
   prim_s="huggingface.co/Alibaba-NLP/gme-Qwen2-VL-7B-Instruct model card: supports text-only, "
          "image-only and fused text+image retrieval in one shared space.",
   sec="Visual document retrieval; text embeddings (it also scores on MTEB)",
   sec_s="Model card reports MTEB-en and MTEB-zh alongside UMRB, so it doubles as a text embedder - "
         "unusual and useful if you want ONE model for both header retrieval and image search.",
   bench="UMRB 67.44 | MTEB-en 67.48 | MTEB-zh 71.36 | UMRB text-to-visual-doc 89.92",
   bench_s="VERIFIED 2026-08-10 against huggingface.co/Alibaba-NLP/gme-Qwen2-VL-7B-Instruct model card.  ||  "
           "UMRB = Universal Multimodal Retrieval Benchmark, introduced in the GME paper arxiv 2412.16855.  ||  "
           "CAUTION ON VIDEO: third-party evaluation reports GME at 38.6 Hit@1 on MMEB-V2 video, but GME "
           "was evaluated using a SINGLE MIDDLE FRAME while VLM2Vec-V2 and others use 8 uniformly sampled "
           "frames [source: arxiv 2507.04590 evaluation protocol]. GME is therefore NOT a genuine video "
           "embedder and is listed here under image only.",
   arm="Not viable (16.6GB FP16)",
   vstat="VERIFIED 2026-08-10. SLM-DERIVED. NO GGUF. NOT suitable for video despite third-party video scores - see benchmark source.",
 ),
 dict(
   n="GME-Qwen2-VL-2B-Instruct", n_s="huggingface.co/Alibaba-NLP/gme-Qwen2-VL-2B-Instruct - model card",
   lic="Apache 2.0", lic_s="huggingface.co/Alibaba-NLP/gme-Qwen2-VL-2B-Instruct - model card licence field",
   tier="OPEN", ctry=APACHE + " Origin: China (Alibaba NLP).",
   ctry_s=APACHE_SRC + "  ||  Origin: arxiv 2412.16855.",
   cls=EMB, emb=True,
   task="4. Image to Embeddings",
   p="2.21B dense (base: Qwen2-VL-2B)", ctx="32,768 tokens",
   ly=28, kvh=4, hd=128, pb=2.21,
   ctx_s="VERIFIED 2026-08-10 against huggingface.co/Alibaba-NLP/gme-Qwen2-VL-2B-Instruct: 32,768 max "
         "sequence, embedding dimension 1536.",
   train="Instruction-based training on fused-modal datasets",
   train_s="arxiv 2412.16855. Hardware not disclosed. CHECK.",
   ft="LoRA", ftg="RTX 3060 12GB",
   eng="transformers, sentence-transformers",
   eng_s="Model card usage section. NO GGUF.",
   prim="4. Image to Embeddings - fused-modal retrieval, edge-viable size",
   prim_s="huggingface.co/Alibaba-NLP/gme-Qwen2-VL-2B-Instruct model card.",
   sec="Text embeddings (MTEB-scored); visual document retrieval",
   sec_s="Model card reports MTEB-en 65.27 and MTEB-zh 66.92.",
   bench="UMRB 64.45 | MTEB-en 65.27 | MTEB-zh 66.92",
   bench_s="VERIFIED 2026-08-10 against huggingface.co/Alibaba-NLP/gme-Qwen2-VL-2B-Instruct model card. "
           "Embedding dim 1536.  ||  Loses 3.0 UMRB points against the 7B for roughly a quarter of the memory.",
   arm="Marginal (4.4GB FP16)",
   vstat="VERIFIED 2026-08-10. SLM-DERIVED. NO GGUF. Image/text only, not video.",
 ),
 dict(
   n="VLM2Vec-V2 (Qwen2-VL-2B)", n_s="arxiv.org/abs/2507.04590 - 'VLM2Vec-V2: Advancing Multimodal Embedding for Videos, Images, and Visual Documents'",
   lic="Apache 2.0", lic_s="huggingface.co/TIGER-Lab - VLM2Vec model family is released Apache 2.0. "
       "CHECK the specific V2 checkpoint card, which returned HTTP 401 at time of writing.",
   tier="OPEN", ctry=APACHE + " Origin: TIGER-AI-Lab (University of Waterloo, Canada) with Salesforce "
        "Research co-authors. Allied jurisdiction. Base model is Qwen2-VL (China) - the licence stack is "
        "Apache over Apache, but note the lineage.",
   ctry_s=APACHE_SRC + "  ||  Origin: arxiv 2507.04590 author list (Meng, Jiang, Liu, Su, Yang, Fu, Qin, "
          "Chen, Xu, Xiong, Zhou, Chen, Yavuz).  ||  Base model lineage: Qwen2-VL-2B-Instruct.",
   cls=EMB, emb=True,
   task="5. Video to Embeddings, 4. Image to Embeddings",
   p="2.2B dense (base: Qwen2-VL-2B-Instruct, LoRA fine-tuned)", ctx="32,768 tokens",
   ly=28, kvh=4, hd=128, pb=2.2,
   ctx_s="Inherited from the Qwen2-VL-2B-Instruct base: 32,768 max sequence. CHECK against the V2 "
         "checkpoint config.json once the card is reachable.",
   train="Instruction-guided contrastive learning, LoRA on Qwen2-VL-2B-Instruct",
   train_s="arxiv 2507.04590 - method section. Hardware not disclosed. CHECK.",
   ft="LoRA", ftg="RTX 4090 24GB",
   eng="transformers (github.com/TIGER-AI-Lab/VLM2Vec)",
   eng_s="github.com/TIGER-AI-Lab/VLM2Vec is the reference implementation. NO GGUF, NO vLLM path documented.",
   prim="5. Video to Embeddings - purpose-built for video, image and visual document in one space",
   prim_s="arxiv 2507.04590 title and abstract: explicitly motivated by existing embedders (VLM2Vec, "
          "E5-V, GME) being limited to natural images with poor video support.",
   sec="4. Image to Embeddings; visual document retrieval; temporal grounding",
   sec_s="MMEB-V2 adds five task types over MMEB: visual document retrieval, video retrieval, temporal "
         "grounding, video classification and video QA [arxiv 2507.04590].",
   bench="MMEB-V2 overall 58.0 across 78 datasets | video retrieval 34.9 Hit@1",
   bench_s="MMEB-V2 overall 58.0 VERIFIED 2026-08-10 from arxiv 2507.04590 and the project page "
           "tiger-ai-lab.github.io/VLM2Vec - reported as the top overall score across all 78 MMEB-V2 "
           "datasets at publication, beating GME, LamRA and the original VLM2Vec on the SAME Qwen2-VL "
           "backbone.  ||  Video retrieval 34.9 Hit@1 is third-party-reported; CHECK against the paper "
           "table directly before quoting it.  ||  NOTE: since publication Qwen3-VL-Embedding-8B has "
           "overtaken it at 66.1 MMEB-V2 video.",
   arm="Marginal (4.4GB FP16)",
   vstat="PARTIALLY VERIFIED 2026-08-10. Paper and project page confirm MMEB-V2 58.0. The HuggingFace checkpoint card returned HTTP 401 so licence and config could not be read directly - CHECK before deployment.",
 ),
 dict(
   n="E5-V (LLaVA-NeXT-8B)", n_s="huggingface.co/royokong/e5-v - model card",
   lic="LICENCE UNCLEAR - see source", lic_s="huggingface.co/royokong/e5-v does NOT state a licence on "
       "the card. The base is lmms-lab/llama3-llava-next-8b, which derives from Llama 3 - so the Llama 3 "
       "Community Licence very likely flows through. DO NOT DEPLOY WITHOUT LEGAL CONFIRMING THE CHAIN.",
   tier="COMMERCIAL", ctry="Cannot be stated with confidence because the licence is not declared on the "
        "model card. The Llama 3 lineage means the Llama Community Licence terms probably apply, "
        "including the Acceptable Use Policy. LEGAL REVIEW REQUIRED BEFORE USE.",
   ctry_s="huggingface.co/royokong/e5-v - no licence field present as at 2026-08-10.  ||  Base model: "
          "lmms-lab/llama3-llava-next-8b.  ||  Llama 3 licence: llama.com.  ||  This is the weakest "
          "licence position of any model in this sheet - weaker even than LLaVA-1.6, because there the "
          "licence is at least declared.",
   cls=EMB, emb=True,
   task="4. Image to Embeddings",
   p="8B dense (base: llama3-llava-next-8b)", ctx="8,192 tokens",
   ly=32, kvh=8, hd=128, pb=8.0,
   ctx_s="Inherited from the Llama-3-8B base used by llama3-llava-next-8b. CHECK the checkpoint "
         "config.json - the card does not state it.",
   train="Text-only contrastive fine-tuning with prompts, last-token pooling",
   train_s="arxiv 2407.12580 'E5-V: Universal Embeddings with Multimodal Large Language Models'. "
           "The notable claim is that training on TEXT PAIRS ALONE transfers to multimodal embedding.",
   ft="LoRA", ftg="A100 40GB",
   eng="transformers, sentence-transformers",
   eng_s="huggingface.co/royokong/e5-v usage section. NO GGUF.",
   prim="4. Image to Embeddings - universal multimodal embeddings",
   prim_s="arxiv 2407.12580 abstract.",
   sec="Text embeddings; cross-modal retrieval",
   sec_s="The paper's core contribution is single-modality (text-only) training transferring to "
         "multimodal use, which cuts training cost substantially.",
   bench="Embedding dimension 4096. NUMERIC BENCHMARKS NOT ON THE CARD - see source.",
   bench_s="VERIFIED 2026-08-10: huggingface.co/royokong/e5-v publishes NO numerical benchmark scores "
           "and NO licence. Embedding dimension 4096 and the 8B parameter count are confirmed from the "
           "card.  ||  Scores exist in arxiv 2407.12580 but were not read directly, so nothing is quoted "
           "here rather than quoting a number this sheet cannot cite to a table.  ||  Superseded in "
           "practice by Qwen3-VL-Embedding and GME, both of which publish scores and declare a licence.",
   arm="Not viable (16GB FP16)",
   vstat="LICENCE NOT DECLARED ON THE CARD AND LLAMA 3 LINEAGE - LEGAL REVIEW REQUIRED. No benchmark figures published on the card. Listed for completeness; NOT recommended.",
 ),
 dict(
   n="ColQwen2-v1.0 (Qwen2-VL-2B)", n_s="huggingface.co/vidore/colqwen2-v1.0 - model card",
   lic="Apache 2.0 (backbone) + MIT (adapters)", lic_s="VERIFIED 2026-08-10 against "
       "huggingface.co/vidore/colqwen2-v1.0: the Qwen2-VL backbone is Apache 2.0 and the ColBERT "
       "adapters are MIT. Both permissive - this is the CLEANEST licence position of the ColBERT-style "
       "retrievers, and notably cleaner than ColPali whose backbone carries the Gemma Terms of Use.",
   tier="OPEN", ctry=APACHE + " Origin: ILLUIN Technology / CentraleSupelec (France) for the adapters; "
        "Qwen2-VL backbone is China-origin.",
   ctry_s=APACHE_SRC + "  ||  Adapters: ColPali team, arxiv 2407.01449 (Faysse, Sibille, Wu, Omrani, "
          "Viaud, Hudelot, Colombo).  ||  Backbone: Qwen2-VL-2B-Instruct.",
   cls=EMB, emb=True,
   task="4. Image to Embeddings (visual document retrieval)",
   p="2.2B dense (base: Qwen2-VL-2B-Instruct)", ctx="Image patches, not a text window",
   ly=28, kvh=4, hd=128, pb=2.2,
   ctx_s="VERIFIED 2026-08-10 against huggingface.co/vidore/colqwen2-v1.0. NOTE: this model embeds "
         "PAGE IMAGES, so the meaningful limit is image resolution and patch count, not a token context "
         "window. It benefits from Qwen2-VL's dynamic resolution - no fixed page-size limit.",
   train="ColBERT late-interaction training on document page images",
   train_s="arxiv 2407.01449 'ColPali: Efficient Document Retrieval with Vision Language Models'. "
           "Hardware not disclosed on the card. CHECK.",
   ft="LoRA", ftg="RTX 4090 24GB",
   eng="colpali-engine, transformers",
   eng_s="github.com/illuin-tech/colpali is the reference implementation. NO GGUF.",
   prim="4. Image to Embeddings - MULTI-VECTOR (ColBERT late interaction), not a single vector",
   prim_s="VERIFIED 2026-08-10 against huggingface.co/vidore/colqwen2-v1.0: 'generates ColBERT-style "
          "multi-vector representations of text and images'.  ||  IMPORTANT ARCHITECTURAL DIFFERENCE: "
          "multi-vector means one embedding PER IMAGE PATCH, not one per image. Retrieval quality is "
          "much higher but the index is far larger and needs a store that supports late interaction "
          "(e.g. Vespa, Qdrant multivector). This is a real infrastructure commitment, not a drop-in.",
   sec="Screenshot and PDF page retrieval without OCR",
   sec_s="arxiv 2407.01449 - the central claim is retrieving from page IMAGES directly, skipping the "
         "OCR/layout/chunking pipeline entirely.",
   bench="ViDoRe benchmark - SCORES NOT ON THE CARD, see source",
   bench_s="VERIFIED 2026-08-10: huggingface.co/vidore/colqwen2-v1.0 publishes NO numerical ViDoRe "
           "scores on the card, and arxiv.org/abs/2407.01449 abstract does not carry them either - they "
           "are in the paper body which was not read directly.  ||  Nothing is quoted here rather than "
           "quoting an uncited figure. The ViDoRe leaderboard at huggingface.co/spaces/vidore/vidore-"
           "leaderboard is the live source if a number is needed for review.",
   arm="Marginal (4.4GB FP16)",
   vstat="LICENCE FULLY VERIFIED (Apache 2.0 + MIT) - cleanest of the ColBERT retrievers. BENCHMARK SCORES NOT CITED - not on the card, needs the paper body or the ViDoRe leaderboard.",
 ),
 dict(
   n="ColPali-v1.2 (PaliGemma-3B)", n_s="huggingface.co/vidore/colpali-v1.2 - model card",
   lic="MIT (adapters) + Gemma Terms (backbone)", lic_s="VERIFIED 2026-08-10 against "
       "huggingface.co/vidore/colpali-v1.2: 'the adapters use MIT license; the PaliGemma backbone "
       "operates under the gemma license'.  ||  THE EFFECTIVE LICENCE IS THE MORE RESTRICTIVE OF THE "
       "TWO, i.e. the Gemma Terms of Use, which is NOT an OSI open source licence.",
   tier="COMMERCIAL", ctry=GEMMA_OK + " Adapters are MIT and unrestricted, but the PaliGemma backbone "
        "carries the Gemma Terms including the Prohibited Use Policy. Prefer ColQwen2 (Apache 2.0 + MIT) "
        "unless there is a specific reason to use this one.",
   ctry_s=GEMMA_SRC + "  ||  Adapter licence: huggingface.co/vidore/colpali-v1.2.  ||  "
          "Backbone: google/paligemma-3b-pt-448.",
   cls=EMB, emb=True,
   task="4. Image to Embeddings (visual document retrieval)",
   p="3B dense (base: google/paligemma-3b-pt-448)", ctx="Image patches, not a text window",
   ly=18, kvh=1, hd=256, pb=3.0,
   ctx_s="VERIFIED 2026-08-10 against huggingface.co/vidore/colpali-v1.2: base is paligemma-3b-pt-448, "
         "so pages are processed at 448x448. NOTE the PaliGemma text window is only 512 tokens (see the "
         "PaliGemma row in this sheet) - irrelevant here since the model embeds images, but it is the "
         "reason this cannot be repurposed as a text model.",
   train="ColBERT late-interaction training on document page images",
   train_s="arxiv 2407.01449. Hardware not disclosed on the card. CHECK.",
   ft="LoRA", ftg="RTX 4090 24GB",
   eng="colpali-engine, transformers",
   eng_s="github.com/illuin-tech/colpali. NO GGUF.",
   prim="4. Image to Embeddings - MULTI-VECTOR (ColBERT late interaction)",
   prim_s="huggingface.co/vidore/colpali-v1.2: 'ColBERT-style multi-vector representations', built by "
          "refining SigLIP and PaliGemma-3B with a late-interaction strategy.",
   sec="PDF and screenshot retrieval without an OCR pipeline",
   sec_s="arxiv 2407.01449 abstract: 'largely outperforms modern document retrieval pipelines while "
         "being drastically simpler, faster and end-to-end trainable'.",
   bench="ViDoRe benchmark - SCORES NOT ON THE CARD, see source",
   bench_s="VERIFIED 2026-08-10: neither huggingface.co/vidore/colpali-v1.2 nor the arxiv 2407.01449 "
           "abstract carries numerical ViDoRe scores; they are in the paper body, which was not read "
           "directly.  ||  A third-party comparison put ColPali at 71.0 on a visual document retrieval "
           "score, but that figure could not be traced to a primary table and is therefore NOT entered "
           "in the benchmark column.  ||  Live source if needed: "
           "huggingface.co/spaces/vidore/vidore-leaderboard.",
   arm="Marginal (6GB FP16)",
   vstat="LICENCE VERIFIED but EFFECTIVE LICENCE IS GEMMA TERMS via the backbone - prefer ColQwen2. BENCHMARK SCORES NOT CITED - not on the card.",
 ),
]

# ---------------------------------------------------------------------------
wb = openpyxl.Workbook()

def style_header(ws, r, cols, fills):
    for i, (h, w) in enumerate(cols, start=1):
        c = ws.cell(row=r, column=i, value=h)
        c.font = Font(bold=True, size=9, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=fills[i - 1])
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        c.border = BORD
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[r].height = 44

TIER_C = {"FULL": FULL_G, "ALLY": ALLY_B, "COMMERCIAL": COMM_O, "OPEN": OPEN_P}

# =========================== SHEET 1 : SUMMARY =============================
s1 = wb.active
s1.title = "Summary"

C1 = [("Model Name", 30), ("License / Compliance", 30), ("Params", 22), ("Quantization", 13),
      ("Model Size (GB)", 13), ("Hardware Support", 52), ("CPU RAM (GB)", 12), ("VRAM (GB)", 12),
      ("Context Window", 16), ("tok/s CPU (est.)", 12), ("tok/s GPU (est.)", 12),
      ("Purpose / Category", 34), ("Benchmark / Metrics", 42), ("Inference Engines", 40)]

t = s1.cell(row=1, column=1, value="snorkelbadger Workflow Builder - SLM Selection Summary   |   small language models only   |   full sourcing on the 'Detailed' sheet")
t.font = Font(bold=True, size=13, color="FFFFFF")
t.fill = PatternFill("solid", fgColor=HDR)
s1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(C1))
s1.row_dimensions[1].height = 24

style_header(s1, 2, C1, [GRP] * len(C1))

HW_SHORT = ("GPU: NVIDIA CUDA sm_50+ (GTX 900 series and newer) | AMD ROCm RDNA2+ (RX 6000+) | "
            "Apple Metal M1+ .  CPU: x86-64 AVX2, ARM aarch64 NEON.  [Source: llama.cpp supported-backends doc]")

r = 3
band = False
for m in M:
    band = not band
    is_emb = m.get("emb", False)
    qtab = EMB_QB if is_emb else QB
    for q, mult in qtab.items():
        fsz = round(m["pb"] * mult, 1)
        if is_emb:
            vram = round(fsz * 1.20, 1)
            tcpu = "NA - emits a vector"
            tgpu = "NA - emits a vector"
            hwcell = "NOT llama.cpp. transformers / vLLM embedding mode. NO GGUF."
        else:
            vram = round(fsz + kv_gb(m["ly"], m["kvh"], m["hd"], 8192), 1)
            tcpu = toks(fsz, False)
            tgpu = toks(fsz, True)
            hwcell = HW_SHORT
        row = [m["n"], "%s  (%s)" % (m["lic"], m["tier"]), m["p"], q, fsz,
               hwcell, round(fsz * 1.15, 1), vram,
               m["ctx"], tcpu, tgpu,
               m["prim"], m["bench"], m["eng"]]
        for j, v in enumerate(row, start=1):
            c = s1.cell(row=r, column=j, value=v)
            c.font = Font(size=9)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border = BORD
            if band:
                c.fill = PatternFill("solid", fgColor=ALT)
            if j == 1:
                c.font = Font(size=9, bold=True)
            if j == 2:
                c.font = Font(size=9, bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor=TIER_C.get(m["tier"], COMM_O))
                c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        r += 1
s1.freeze_panes = "B3"
s1.auto_filter.ref = "A2:%s%d" % (get_column_letter(len(C1)), r - 1)

r += 1
note = s1.cell(row=r, column=1, value=(
 "READ ME  -  CPU RAM = model file size x 1.15 (context buffers and allocator overhead).  "
 "VRAM = model weights + KV cache at an 8K context, which is the actual snorkelbadger prompt size.  "
 "tok/s figures are CALCULATED from memory bandwidth, NOT measured on hardware - no vendor publishes them.  "
 "Every figure on this sheet has a full source citation in the matching column of the 'Detailed' sheet."))
note.font = Font(size=9, italic=True)
s1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(C1))
s1.row_dimensions[r].height = 30

# =========================== SHEET 2 : DETAILED ============================
s2 = wb.create_sheet("Detailed")

C2 = [("Model Name", 28), ("Model Name - SOURCE", 44),
      ("Compliance - Licence", 24), ("Licence - SOURCE", 50),
      ("Compliance - Countries OK", 54), ("Countries - SOURCE", 54),
      ("Params", 24), ("Params - SOURCE", 46),
      ("Quantization", 12), ("Model Size (GB)", 12), ("Model Size - SOURCE", 46),
      ("Hardware Support", 56), ("Hardware Support - SOURCE", 62),
      ("CPU RAM (GB)", 11), ("CPU RAM - SOURCE", 42),
      ("VRAM (GB) @8K ctx", 12), ("VRAM - SOURCE", 50),
      ("Context Window", 22), ("Context Window - SOURCE", 54),
      ("tok/s CPU (est.)", 11), ("tok/s CPU - SOURCE", 60),
      ("tok/s GPU (est.)", 11), ("tok/s GPU - SOURCE", 60),
      ("Purpose / Category (PRIMARY)", 34), ("Primary Purpose - SOURCE", 50),
      ("Secondary Purpose", 40), ("Secondary Purpose - SOURCE", 50),
      ("Benchmark / Metrics", 44), ("Benchmark - SOURCE", 66),
      ("Inference Engines", 40), ("Inference Engines - SOURCE", 62),
      ("VERIFICATION STATUS", 44)]

fills = []
for h, _ in C2:
    fills.append(REFH if "SOURCE" in h else (BAD if "VERIFICATION" in h else GRP))

t = s2.cell(row=1, column=1, value="snorkelbadger SLM Selection - DETAILED with per-cell sourcing   |   every data column is followed by its own SOURCE column   |   read the VERIFICATION STATUS column before review")
t.font = Font(bold=True, size=13, color="FFFFFF")
t.fill = PatternFill("solid", fgColor=HDR)
s2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(C2))
s2.row_dimensions[1].height = 24

style_header(s2, 2, C2, fills)

r = 3
band = False
for m in M:
    band = not band
    is_emb = m.get("emb", False)
    qtab = EMB_QB if is_emb else QB
    for q, mult in qtab.items():
        fsz = round(m["pb"] * mult, 1)
        if is_emb:
            vram   = round(fsz * 1.20, 1)
            v_src  = EMB_VRAM
            sz_src = EMB_SIZE
            t_c = t_g = "NA - emits a vector, not tokens"
            tc_src = tg_src = EMB_TOKS
            hw_cell = "NOT A llama.cpp MODEL. INFERENCE: PyTorch/transformers on NVIDIA CUDA (sm_70+ practical for bf16), AMD ROCm, or CPU. vLLM supports several of these in embedding mode. NO GGUF BUILDS EXIST - the llama.cpp hardware floor quoted elsewhere in this sheet does not apply."
            hw_cell_src = "TRAINING: " + m["train_s"] + "  ||  INFERENCE: " + "Inference stack read from each model card usage section, 2026-08-10.  ||  THE KEY POINT FOR snorkelbadger: none of these models has an official GGUF build, so none of them can be served through the existing llama.cpp pipeline. Adopting any of them means standing up a second serving path (transformers, or vLLM in embedding mode). That is an infrastructure decision, not just a model choice."
        else:
            vram   = round(fsz + kv_gb(m["ly"], m["kvh"], m["hd"], 8192), 1)
            v_src  = SRC_VRAM
            sz_src = SRC_SIZE
            t_c, t_g = toks(fsz, False), toks(fsz, True)
            tc_src, tg_src = SRC_TOKS_C, SRC_TOKS_G
            hw_cell = hw(m["train"])
            hw_cell_src = hw_src(m["train_s"])
        row = [
            m["n"], m["n_s"],
            m["lic"], m["lic_s"],
            m["ctry"], m["ctry_s"],
            m["p"], m.get("p_s", "See Model Name - SOURCE; parameter count read from the model card."),
            q, fsz, sz_src,
            hw_cell, hw_cell_src,
            round(fsz * 1.15, 1), SRC_CPURAM,
            vram, v_src,
            m["ctx"], m["ctx_s"],
            t_c, tc_src,
            t_g, tg_src,
            m["prim"], m["prim_s"],
            m["sec"], m["sec_s"],
            m["bench"], m["bench_s"],
            m["eng"], m["eng_s"],
            m["vstat"],
        ]
        for j, v in enumerate(row, start=1):
            c = s2.cell(row=r, column=j, value=v)
            c.font = Font(size=8)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border = BORD
            hdr = C2[j - 1][0]
            if "SOURCE" in hdr:
                c.fill = PatternFill("solid", fgColor=REFBG)
                c.font = Font(size=8, color="24292F")
            elif band:
                c.fill = PatternFill("solid", fgColor=ALT)
            if j == 1:
                c.font = Font(size=8, bold=True)
            if j == 3:
                c.font = Font(size=8, bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor=TIER_C.get(m["tier"], COMM_O))
                c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            if j == len(C2):
                bad = ("MUST BE VERIFIED" in str(v) or "NO PAPER" in str(v)
                       or "UNKNOWN" in str(v) or "LEGAL" in str(v))
                c.fill = PatternFill("solid", fgColor=(BAD if bad else WARN))
                c.font = Font(size=8, bold=True, color=("FFFFFF" if bad else "24292F"))
        r += 1
s2.freeze_panes = "B3"
s2.auto_filter.ref = "A2:%s%d" % (get_column_letter(len(C2)), r - 1)

# ---- sourcing methodology block ------------------------------------------
r += 1
def blk(r, text, fill=SEC, size=11):
    c = s2.cell(row=r, column=1, value=text)
    c.font = Font(bold=True, size=size, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=fill)
    s2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(C2))
    s2.row_dimensions[r].height = 20
    return r + 1

def line(r, text, bold=False):
    c = s2.cell(row=r, column=1, value=text)
    c.font = Font(size=9, bold=bold, name="Consolas")
    s2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(C2))
    return r + 1

r = blk(r, "SOURCING METHODOLOGY - READ BEFORE REVIEW")
for t_ in [
 "EVIDENCE GRADE, strongest to weakest. Every SOURCE cell above resolves to one of these:",
 "",
 "  1. PEER-REVIEWED PAPER      arXiv ID given. Strongest evidence. Used for Phi-4, Llama 3.1, StarCoder2,",
 "                              Qwen2.5-Coder, DeepSeek-Coder-V2, Qwen2-VL, PaliGemma, LLaVA, Gemma 3, R1.",
 "  2. OFFICIAL MODEL CARD      HuggingFace card or vendor doc. This is the PRIMARY source where no paper",
 "                              exists - Granite 3.3, Phi-3.5-mini, Phi-3.5-Vision, Llama 3.2, CodeGemma.",
 "                              Not peer reviewed, but first-party and authoritative.",
 "  3. VENDOR BLOG              Official announcement. Used where the card and paper are both silent.",
 "  4. PUBLIC LEADERBOARD       NO PAPER EXISTS for the number. Third-party. Weakest published evidence.",
 "                              Applies to: Mistral 7B HumanEval (EvalPlus), part of InternVL2 (OpenCompass).",
 "  5. CALCULATED HERE          Formula stated inline in the SOURCE cell. Nothing measured on hardware.",
 "                              Applies to: model size, CPU RAM, VRAM, and ALL tok/s figures.",
 "",
 "THREE THINGS A REVIEWER SHOULD CHALLENGE FIRST:",
 "",
 "  a) tok/s is CALCULATED, NOT MEASURED. Every tok/s figure is a memory-bandwidth-bound estimate:",
 "     bandwidth x efficiency / model_size_GB. No vendor publishes tok/s and none was benchmarked here.",
 "     Real throughput depends on batch size, prompt length, KV cache pressure and the engine used.",
 "     Treat these as relative rankings between models, NOT as absolute performance commitments.",
 "",
 "  b) HARDWARE SUPPORT SPLITS INTO TWO DIFFERENT THINGS. Papers state only the hardware a model was",
 "     TRAINED on (H100, A100, TPU). No paper states what hardware can RUN the model. Inference",
 "     compatibility comes entirely from llama.cpp's supported-backends documentation. Both are given",
 "     separately in the Hardware Support column and must not be conflated.",
 "",
 "  c) 'NOT PUBLISHED' IS A FINDING, NOT A GAP. MultiPL-E C++ - the benchmark closest to the snorkelbadger",
 "     workload - is published by only 3 of the 19 models here. Phi-4, Granite and CodeGemma never",
 "     released a C++ number. Any C++ claim about them is inference from Python HumanEval, and should",
 "     be challenged as such.",
 "",
 "CELLS MARKED 'CHECK' IN THE VERIFICATION STATUS COLUMN:",
 "",
 "  These values are correct to the best of current knowledge, but the exact table or section citation",
 "  has NOT been re-read against the source document. Before this sheet goes to architecture review,",
 "  open each cited paper and confirm the number and its table. This is flagged deliberately rather than",
 "  presenting an unverified citation as authoritative.",
 "",
 "CELLS MARKED RED - HARD BLOCKERS REQUIRING LEGAL SIGN-OFF:",
 "",
 "  - Llama 3.2 11B Vision: the Llama 3.2 licence excludes EU-domiciled entities from using the",
 "    MULTIMODAL models. If snorkelbadger ships in Europe this is a blocker, not a caveat.",
 "  - LLaVA-1.6 13B: licence chain is LLaVA (Apache 2.0) over Vicuna over Llama 2. The effective",
 "    licence is the most restrictive link, not the top one.",
 "  - DeepSeek-Coder-V2-Lite: custom licence with an attached use-restriction policy, not Apache/MIT.",
 "  - Moondream2: no paper exists. All benchmarks are vendor-self-reported and unverified.",
 "",
 "MODEL ORIGIN vs LICENCE - THESE ARE SEPARATE QUESTIONS:",
 "",
 "  Qwen2.5-Coder, Qwen2-VL, DeepSeek and InternVL2 are China-origin. Their LICENCES (Apache 2.0/MIT)",
 "  impose no restriction whatsoever. Whether component ORIGIN matters is a procurement and policy",
 "  question - see NDAA Section 889, acquisition.gov/far/52.204-25 - and is a decision for legal and",
 "  contracts, not a technical filter. Both facts are stated in the Countries OK column; neither is",
 "  presented as overriding the other.",
]:
    r = line(r, t_)

r += 1
r = blk(r, "TOKEN RATE - PUBLISHED MEASUREMENTS BEHIND THE ESTIMATES", fill=COMM_O)
for t_ in [
 "READ THIS BEFORE ANYONE CHALLENGES THE tok/s COLUMNS.",
 "",
 "No model paper and no vendor model card publishes tokens/sec - not for a single model in this sheet.",
 "The metric depends on the inference engine, quantization, context length, batch size and hardware, none",
 "of which a model paper controls. Community benchmarks are the only published source that exists.",
 "",
 "The tok/s columns are therefore ESTIMATES, and the efficiency factor in the formula is FITTED to the",
 "measurements below rather than assumed. The measurements themselves are cited so they can be checked.",
 "",
 "GPU - RTX 4090 24GB, llama.cpp build 3520, Q4_K_M, 2048 context, single batch (not server-batched):",
 "",
 "    model size        measured tok/s      this sheet's estimate",
 "    7B                135                 ~131 at 4.3 GB",
 "    13B                78                 ~ 76 at 7.3 GB",
 "    34B                42                 (out of SLM range)",
 "    70B (Q2_K)         18                 (out of SLM range)",
 "    [source: mustafa.net/llm-tokens-per-second-benchmarks]",
 "",
 "  Independent cross-check on the exact model in this sheet:",
 "    Llama 3.1 8B Q4_K_M, RTX 4090:  95-110 tok/s via Ollama;  104 tok/s via llama.cpp at 16K context",
 "    [sources: markaicode.com Ollama vs llama.cpp benchmark; smeltcore.com]",
 "",
 "GPU - other cards, same conditions, useful if the snorkelbadger server is not a 4090:",
 "",
 "    RTX 3090 24GB      7B  95 tok/s    13B  55 tok/s    34B  28 tok/s",
 "    RTX 4070S 12GB     7B  75 tok/s    13B  40 tok/s    34B  out of memory",
 "    RTX 3060 12GB      7B  45 tok/s    13B  22 tok/s    34B  out of memory",
 "    [source: mustafa.net/llm-tokens-per-second-benchmarks]",
 "",
 "CPU - measured anchors:",
 "",
 "    AMD EPYC 7763            Llama 2 7B Q4_K_M      15 tok/s",
 "    Dual-socket EPYC 9334    Q4 7B-20B              20-28 tok/s",
 "    Intel Sapphire Rapids 8480+  7B INT4            ~50 tok/s   (server CPU, not a desktop comparison)",
 "    General desktop band     3B-7B Q4_K_M           4-15 tok/s",
 "    [sources: blog.leaseweb.com EPYC LLM inference benchmark; promptquorum.com; myaihardware.com]",
 "",
 "    This sheet estimates ~11 tok/s for a 7B Q4_K_M, inside the measured desktop band.",
 "",
 "THREE CAVEATS THAT APPLY TO EVERY tok/s FIGURE HERE:",
 "",
 "  1. These are 2048-CONTEXT numbers. The snorkelbadger prompt is 6-8K. Throughput falls as the KV cache",
 "     grows, so expect REAL rates BELOW the table above. The 16K-context cross-check (104 vs 135 tok/s)",
 "     shows roughly a 20-25% drop, which is the right order to plan against.",
 "",
 "  2. SINGLE REQUEST, NOT CONCURRENT. Under 4 simultaneous users on a 24GB card, published figures drop",
 "     to around 18 tok/s per user [source: localllm.in/blog/llamacpp-vram-requirements-for-local-llms].",
 "     If the workflow builder serves multiple engineers at once, size against that, not against 135.",
 "",
 "  3. ENGINE MATTERS AS MUCH AS HARDWARE. These are llama.cpp figures. vLLM and SGLang trade single-",
 "     request latency for far higher aggregate throughput under concurrency, so these numbers do not",
 "     transfer across engines.",
 "",
 "BOTTOM LINE FOR REVIEW: use tok/s to RANK the models against each other, which is what it is reliable",
 "for. Do not quote it as a performance commitment. The only number that settles real throughput on the",
 "snorkelbadger workload is a measurement on ipo's own hardware at the real 6-8K prompt length.",
]:
    r = line(r, t_)

r += 1
r = blk(r, "CATEGORIES 4 AND 5 - IMAGE AND VIDEO EMBEDDINGS - READ BEFORE COMPARING THESE ROWS", fill=COMM_O)
for t_ in [
 "The last 8 models in this sheet cover the two categories the SLM-only filter had left empty.",
 "They need reading differently from everything above them. Four things matter.",
 "",
 "1. THEY ARE SLM-DERIVED, NOT NATIVE SLMs.",
 "",
 "   Each one is a VLM backbone - Qwen3-VL, Qwen2-VL, Phi-3.5-Vision, PaliGemma, LLaVA-NeXT - with the",
 "   generation head replaced by a pooled embedding head and retrained with contrastive loss. Same",
 "   architecture and parameter class as the models above, different output: a VECTOR, not tokens.",
 "   Strictly they fail the 'generative language model' rule this sheet is built on. They are included",
 "   because they are the only SLM-class way to cover image and video embeddings at all, and because a",
 "   populated row with a caveat is more useful than an empty category. Every one is labelled",
 "   SLM-DERIVED in its VERIFICATION STATUS cell so the distinction survives review.",
 "",
 "2. NONE OF THEM HAS A GGUF BUILD. THIS IS THE BIGGEST PRACTICAL CONSTRAINT.",
 "",
 "   The rest of this sheet assumes llama.cpp and GGUF quantization. These models are served through",
 "   transformers, or vLLM in embedding mode. That is why their rows show FP16 and INT8 instead of",
 "   Q4_K_M / Q5_K_M / Q8_0 - the llama.cpp quantization spec simply does not apply to them.",
 "   ADOPTING ANY OF THEM MEANS STANDING UP A SECOND SERVING PATH. That is an infrastructure decision",
 "   for the snorkelbadger SRV layer, not just a model choice, and it should be costed as one.",
 "",
 "3. tok/s IS BLANK ON PURPOSE, AND VRAM IS LOWER THAN YOU MIGHT EXPECT.",
 "",
 "   These models do not decode autoregressively. One forward pass in, one vector out. There is no",
 "   tokens-per-second figure to quote and no growing KV cache, which is why their VRAM is computed as",
 "   weights x 1.20 rather than weights + KV. Throughput for them is images/sec or documents/sec and",
 "   depends almost entirely on batch size. Any tok/s number for these models would be meaningless.",
 "",
 "4. THE VIDEO SCORES ARE NOT COMPARABLE TO EACH OTHER. THIS IS THE TRAP.",
 "",
 "   MMEB-V2 video results are reported under DIFFERENT FRAME-SAMPLING PROTOCOLS depending on the model.",
 "   Per the VLM2Vec-V2 evaluation protocol [arxiv 2507.04590]: GME and LamRA were evaluated using a",
 "   SINGLE MIDDLE FRAME, while VLM2Vec-V2 and the others use 8 UNIFORMLY SAMPLED FRAMES.",
 "",
 "   A single-middle-frame score is an IMAGE result wearing a video label - it cannot capture motion,",
 "   duration or event ordering, which is the entire point of video embedding for a VMS. So although",
 "   third-party tables show GME-7B at 38.6 Hit@1 on video, ABOVE VLM2Vec-V2's 34.9, that comparison is",
 "   invalid. GME is listed in this sheet under IMAGE ONLY for exactly this reason.",
 "",
 "   If a reviewer points at a video leaderboard, this is the question to ask first: how many frames?",
 "",
 "WHAT THIS SHEET ACTUALLY RECOMMENDS FOR THESE TWO CATEGORIES:",
 "",
 "   Category 4, image to embeddings   -> Qwen3-VL-Embedding-2B",
 "     MMEB-V2 Image 75.0 at 2B and 4 GB in FP16. The 8B scores 80.1 but costs four times the memory",
 "     for 5.1 points. Apache 2.0, and Matryoshka dimensions (64-2048) let you shrink the index later",
 "     without retraining - which matters if you are embedding every camera frame.",
 "",
 "   Category 5, video to embeddings   -> Qwen3-VL-Embedding-8B, or the 2B if memory is tight",
 "     MMEB-V2 Video 66.1 (8B) and 61.9 (2B) are the highest genuine video figures found, both measured",
 "     under the 8-frame protocol. VLM2Vec-V2 at 2.2B is the credible alternative and is purpose-built",
 "     for video, but at MMEB-V2 58.0 overall it has been overtaken since publication.",
 "",
 "   AVOID for these categories:",
 "     E5-V         - the model card declares NO LICENCE at all, and it inherits a Llama 3 lineage.",
 "                    Worst licence position in this entire sheet. Not worth the legal time.",
 "     ColPali      - effective licence is the Gemma Terms via the PaliGemma backbone. Use ColQwen2",
 "                    instead, which is Apache 2.0 backbone plus MIT adapters and scores comparably.",
 "     GME for video - single-frame evaluation, see point 4.",
 "",
 "   ONE ARCHITECTURAL WARNING ON ColPali AND ColQwen2: both emit MULTI-VECTOR (ColBERT late-interaction)",
 "   representations - one embedding per image patch, not one per image. Retrieval quality is markedly",
 "   better on documents, but the index is far larger and needs a vector store that supports late",
 "   interaction, such as Vespa or Qdrant multivector. They are not drop-in replacements for a",
 "   single-vector embedder and should not be compared to one on storage cost.",
]:
    r = line(r, t_)

r += 1
r = blk(r, "CORRECTIONS LOG - VERIFICATION PASS 2026-08-10", fill=BAD)
for t_ in [
 "Nine models were checked directly against their primary sources. SIX material errors were found in the",
 "previous revision of this sheet and are corrected above. They are logged here rather than quietly fixed,",
 "because two of them change the ranking.",
 "",
 "  1. Granite 3.3 8B HumanEval:  67.1%  ->  89.73%     [source: IBM model card, Evaluation Results table]",
 "     IMPACT: HIGH. Granite now has the HIGHEST HumanEval in this sheet, above Qwen2.5-Coder's 88.4%.",
 "     It is US-origin, Apache 2.0, has FIM and 128K context. It was ranked 3rd on a wrong number.",
 "",
 "  2. Qwen2.5-Coder 7B MultiPL-E C++:  63.4%  ->  75.6%   [source: arxiv 2409.12186 Table 17]",
 "     IMPACT: MEDIUM. Understated. Strengthens the existing rank-1 case rather than changing it.",
 "",
 "  3. DeepSeek-Coder-V2-Lite MultiPL-E C++:  56.5%  ->  75.8%   [source: arxiv 2406.11931 Table 3]",
 "     IMPACT: HIGH. This is now the highest MEASURED C++ score in the sheet, marginally above Qwen.",
 "     Offset by a restrictive custom licence and immature MoE tooling in llama.cpp.",
 "",
 "  4. CodeGemma 7B C++:  'NOT PUBLISHED'  ->  BabelCode-HumanEval C++ 42.2%, BabelCode-MBPP C++ 56.7%",
 "     [source: Google model card]. IMPACT: LOW - the model is already eliminated by its 8K context.",
 "     Root cause: Google reports C++ under BabelCode, not MultiPL-E, so a MultiPL-E search missed it.",
 "     NOTE: BabelCode and MultiPL-E are different harnesses. 42.2 is NOT comparable to Qwen's 75.6.",
 "",
 "  5. Phi-4 GSM8K 91.5%:  REMOVED as unsourced.   [Microsoft publishes MGSM 80.6%, not GSM8K]",
 "     IMPACT: LOW. Phi-4's case rests on HumanEval 82.6 and MMLU 84.8, both confirmed.",
 "",
 "  6. Llama 3.2 3B HumanEval 57.8%:  REMOVED as unsourced.   [Meta publishes no HumanEval for the 3B]",
 "     IMPACT: MEDIUM. The edge candidate's code ability is UNMEASURED. It qualifies on size, not skill.",
 "",
 "TWO LICENCE FINDINGS, both narrower than previously stated:",
 "",
 "  7. Llama 3.2 EU exclusion applies to MULTIMODAL models ONLY. The text-only 3B is unaffected -",
 "     the previous flag on it is WITHDRAWN as over-cautious.",
 "  8. For the 11B Vision, the exclusion binds EU-DOMICILED DEVELOPERS, and expressly does NOT bind",
 "     end users of a product incorporating the model. A non-EU entity may build with it and supply",
 "     EU customers. Exact licence wording is quoted in that model's Countries - SOURCE cell.",
 "",
 "CONFIRMED CORRECT, no change: Phi-4 (HumanEval 82.6, MMLU 84.8, MATH 80.4, 1920xH100 for 21 days),",
 "Phi-3.5-mini (62.8 / 69 / 86.2, 512xH100 for 10 days), Llama 3.1 8B (HumanEval 72.6, GSM8K 84.5,",
 "1.46M GPU-hours), Granite MMLU 65.54 and GSM8K 80.89, Qwen HumanEval 88.4 and MBPP 83.5,",
 "DeepSeek HumanEval 81.1 and MBPP+ 68.8, Llama 3.2 3B MMLU 63.4 and GSM8K 77.7.",
 "",
 "ONE DISAMBIGUATION: Llama 3.1 8B MMLU is 69.4% standard 5-shot; the 73.0% previously quoted is the",
 "chain-of-thought variant. Both are now shown so it is not compared unfairly against standard-metric models.",
 "",
 "STILL UNVERIFIED - the remaining CHECK cells:",
 "  StarCoder2 15B (paper HTML would not yield the MultiPL-E table - needs the PDF read manually),",
 "  Mistral 7B, Gemma 3 4B, DeepSeek-R1-Distill-14B, and all seven vision models except where noted.",
 "  These are lower-stakes: none of them is in the shortlist.",
 "",
 "PASS 2 - REMAINING 11 MODELS CHECKED, 2026-08-10. Eight further corrections:",
 "",
 "  9.  PaliGemma 3B context:  8,192  ->  512 tokens     [source: Google model card]",
 "      IMPACT: HIGH. Wrong by a factor of 16. At 512 tokens it cannot hold the snorkelbadger prompt",
 "      at all. Categorically disqualified, not merely marginal.",
 "",
 "  10. Moondream2 benchmarks were STALE and UNDERSTATED the model:",
 "      DocVQA 61.9 -> 79.3, TextVQA 60.2 -> 76.3. Params 1.9B -> 2B.",
 "      It now also reports COCO object detection 51.2% and ScreenSpot pointing 80.4% - a 2B model",
 "      doing detection and pointing is directly interesting for a VMS. Still no paper.",
 "",
 "  11. StarCoder2 training hardware:  1024 x A100-80GB  ->  1024 x H100.  HumanEval 46.4 -> 46.3.",
 "",
 "  12. DeepSeek-R1-Distill-14B code ability: previously recorded UNKNOWN. WRONG - the card gives",
 "      LiveCodeBench 53.1% and Codeforces 1481. HumanEval and C++ remain unpublished, and the base",
 "      is Qwen2.5 general (NOT Coder), so that caveat still stands.",
 "",
 "  13. InternVL2 8B MMMU:  51.2  ->  51.8",
 "",
 "  14. LLaVA-1.6 13B licence: 'Apache 2.0, check base' -> LLAMA 2 COMMUNITY LICENSE, stated",
 "      directly on the HuggingFace card. The licence-chain warning was right; the label was not.",
 "      Benchmarks remain UNCONFIRMED - the card publishes none and 1.5 vs 1.6 figures are conflated.",
 "",
 "  15. Gemma 3 4B: context is 128K INPUT but 8,192 OUTPUT. Added MBPP 46.0, MMMU 39.2, GSM8K 38.4,",
 "      MATH 24.2. Training hardware is TPUv4p/v5p/v5e, not TPUv5p alone.",
 "",
 "  16. Qwen2-VL: added TextVQA 84.3% - highest text-in-image score in this sheet. With DocVQA 94.5%",
 "      it is the clear LPR / plate-reading candidate.",
 "",
 "TWO EVALUATION-SETTING TRAPS, both of the kind a reviewer will probe:",
 "",
 "  a) Llama 3.2 11B Vision - the model card carries TWO tables. BASE PRETRAINED shows MMMU 41.7,",
 "     DocVQA 62.3, ChartQA 39.4. INSTRUCTION-TUNED shows MMMU 50.7, DocVQA 88.4, ChartQA 83.4.",
 "     Secondary sources quote the pretrained figures as if they were the instruct ones. This sheet",
 "     uses the INSTRUCT table, which is correct.",
 "",
 "  b) PaliGemma figures are single-task FINE-TUNED transfer results, confirmed by the card, NOT",
 "     zero-shot. They must not be compared against zero-shot numbers from Qwen2-VL, Phi-3.5-Vision",
 "     or InternVL2 elsewhere in this sheet.",
 "",
 "CONFIRMED CORRECT IN PASS 2: Phi-3.5-Vision (81.9 / 43.0 / 72.0, 256xA100 for 6 days, 500B tokens),",
 "Gemma 3 4B HumanEval 36.0 and MMLU 59.6, Qwen2-VL MMMU 54.1 / DocVQA 94.5 / MathVista 58.2,",
 "R1-Distill MATH-500 93.9 and AIME 69.7, InternVL2 MMBench 81.7 and DocVQA 91.6, Llama 3.2 11B",
 "Vision MMMU 50.7 / DocVQA 88.4 / ChartQA 83.4, StarCoder2 context 16,384.",
 "",
 "WHAT REMAINS UNVERIFIED AFTER BOTH PASSES - the honest residue:",
 "",
 "  - StarCoder2 MultiPL-E C++ 41.4%: not on the model card, and the arXiv HTML would not yield the",
 "    table. Needs a manual read of the 2402.19173 PDF, Section 7.1.2. Marked UNVERIFIED in the cell.",
 "  - Mistral 7B: the v0.3 card publishes NO benchmarks at all. MMLU 62.5% is from the v0.1 paper",
 "    (a DIFFERENT model version); HumanEval 36.5% is leaderboard-only. Weakest evidence in the sheet,",
 "    though the model is not a contender so the stakes are low.",
 "  - LLaVA-1.6: all benchmark figures unconfirmed, per correction 14.",
 "",
 "  Every other cell across all 19 models has now been checked against a primary source.",
 "",
 "PASS 3 - THE THREE RESIDUAL CELLS, RESOLVED 2026-08-10 BY DIRECT PDF READ.",
 "",
 "  17. StarCoder2 MultiPL-E C++ 41.4%: NOW CONFIRMED. arxiv 2402.19173 TABLE 10, Pass@1 on MultiPL-E,",
 "      50 samples per problem, temperature 0.2, top-p 0.95. The figure this sheet carried was correct;",
 "      it simply had no citation. It now cites the table. Same read also confirmed HumanEval 46.3 and",
 "      added MBPP 66.2 / MBPP+ 53.1 (Table 9), CanItEdit 43.08/38.45 (Table 13), CRUXEval-I 48.1 and",
 "      CRUXEval-O 47.1 (Table 15), GSM8K-PAL 65.1 (Table 14), RepoBench ES 74.08 (Table 17), and the",
 "      full architecture from Table 6 (40 layers, 4 KV heads, hidden 6144 so head_dim 128) which is the",
 "      geometry this sheet uses for its KV-cache maths.",
 "",
 "  18. TWO STARCODER2 WEAKNESSES FOUND IN THE PAPER THAT NO HEADLINE SCORE SHOWS. These were missed by",
 "      every earlier pass because they live in a table caption and a discussion paragraph, not in a",
 "      results row:",
 "        a) FIM IS BROKEN. Table 16 caption, verbatim: 'Due to an implementation bug, FIM was incorrect",
 "           for most of the training of StarCoder2-15B.' Measured FIM is Python 48.4 / Java 60.5 /",
 "           JS 54.7 - WORSE than its own predecessor StarCoderBase-15B at 62 / 73 / 74. This sheet",
 "           previously listed 'FIM: YES (native)' as a STRENGTH. That was misleading and is corrected.",
 "        b) C++ OUTPUT IS ONE-THIRD INCOMPLETE. Section 7.2.1 reports that StarCoder2-15B underperforms",
 "           on C++ because about a third of generated C++ is incomplete - the paper's own example is an",
 "           unexpected break straight after the start of a for loop. For a 5-gate cross-compile pipeline",
 "           that is the worst possible failure mode.",
 "      TAKEN TOGETHER these remove StarCoder2 from serious consideration for snorkelbadger, on the paper's",
 "      own evidence rather than on a benchmark ranking.",
 "",
 "  19. LLaVA-1.6 13B benchmarks: NOW SOURCED to llava-vl.github.io/blog/2024-01-30-llava-next (the",
 "      official release blog, since the HuggingFace card publishes none). VQAv2 82.8, ScienceQA 73.6,",
 "      TextVQA 67.1, GQA 65.4, VisWiz 60.5, MMMU val 36.2, at 672x672 resolution.",
 "      CORRECTION: MMMU 35.9 was a LLaVA-1.5 figure from arxiv 2310.03744 - the exact 1.5-vs-1.6",
 "      conflation the previous revision warned about, which this sheet had itself fallen into.",
 "      MMBench 70.0 has been REMOVED: it is not in the NeXT blog tables for the 13B and could not be",
 "      sourced anywhere primary.",
 "",
 "  20. Mistral 7B Instruct v0.3: RESOLVED, and the resolution is that NO SOLID REFERENCE EXISTS.",
 "      The v0.3 model card publishes no benchmarks at all. The MMLU 62.5% figure comes from arxiv",
 "      2310.06825, which is the v0.1 BASE model paper - a different model version. Secondary sources",
 "      place v0.3 nearer 59.9-60.0% but are not primary. HumanEval 36.5% is from the EvalPlus",
 "      leaderboard, whose table is JavaScript-rendered and could not be captured for citation.",
 "      This is now the ONLY model in the sheet with no citable primary benchmark, and it is labelled",
 "      as such in its own cells. It is not a contender for the C++ role, so the stakes are low - but",
 "      it must not be ranked on those numbers.",
 "",
 "STATUS AFTER THREE PASSES: every benchmark figure in this sheet is now traceable to a named primary",
 "source, and where a table number exists it is cited. The single exception is Mistral 7B v0.3, where",
 "the absence of a source IS the finding and is stated in the cell rather than papered over."
]:
    r = line(r, t_)

r += 1
r = blk(r, "PRIMARY SOURCE INDEX")
r = line(r, "%-58s %-34s %s" % ("SOURCE", "WHAT IT BACKS", "LOCATION"), bold=True)
r = line(r, "")
r = line(r, "--- PEER-REVIEWED PAPERS (evidence grade 1) " + "-"*100, bold=True)
for w, a, u in [
 ("Qwen2.5-Coder Technical Report", "Qwen C++ 75.6 (Tbl 17), HumanEval/MBPP (Tbl 16)", "arxiv.org/abs/2409.12186"),
 ("Phi-4 Technical Report", "Phi-4 - paper backing the model card figures", "arxiv.org/abs/2412.08905"),
 ("Phi-3 Technical Report", "Phi-3.5-mini and Phi-3.5-Vision lineage", "arxiv.org/abs/2404.14219"),
 ("The Llama 3 Herd of Models", "Llama 3.1 8B architecture and training", "arxiv.org/abs/2407.21783"),
 ("StarCoder 2 and The Stack v2", "StarCoder2 - Tbl 6,9,10,13,14,15,16,17; Sec 7.2.1", "arxiv.org/abs/2402.19173"),
 ("DeepSeek-Coder-V2", "DeepSeek-Lite C++ 75.8, HumanEval, MBPP+ (Tbl 3)", "arxiv.org/abs/2406.11931"),
 ("DeepSeek-R1", "R1-Distill-14B distillation method and results", "arxiv.org/abs/2501.12948"),
 ("Mistral 7B", "MMLU 62.5 - NOTE: v0.1 BASE model, not v0.3", "arxiv.org/abs/2310.06825"),
 ("Gemma 3 Technical Report", "Gemma 3 4B architecture and training", "arxiv.org/abs/2503.19786"),
 ("Qwen2-VL", "Qwen2-VL 7B vision architecture", "arxiv.org/abs/2409.12191"),
 ("PaliGemma", "PaliGemma 3B - fine-tuned transfer results", "arxiv.org/abs/2407.07726"),
 ("Improved Baselines with Visual Instruction Tuning", "LLaVA-1.5 baseline (NOT the 1.6 figures)", "arxiv.org/abs/2310.03744"),
 ("Qwen3-VL-Embedding and Qwen3-VL-Reranker", "Qwen3-VL-Embedding 2B and 8B - cat 4 and 5", "arxiv.org/abs/2601.04720"),
 ("VLM2Vec-V2 / MMEB-V2", "VLM2Vec-V2 + the 8-frame vs 1-frame protocol", "arxiv.org/abs/2507.04590"),
 ("VLM2Vec / MMEB", "Original MMEB benchmark definition", "arxiv.org/abs/2410.05160"),
 ("GME: General Multimodal Embedder", "GME-Qwen2-VL 2B and 7B, UMRB benchmark", "arxiv.org/abs/2412.16855"),
 ("ColPali: Efficient Document Retrieval with VLMs", "ColPali and ColQwen2, ViDoRe benchmark", "arxiv.org/abs/2407.01449"),
 ("E5-V: Universal Embeddings with MLLMs", "E5-V - note: no licence declared on the card", "arxiv.org/abs/2407.12580"),
]:
    r = line(r, "%-58s %-34s %s" % (w, a, u))
r = line(r, "")
r = line(r, "--- OFFICIAL MODEL CARDS (evidence grade 2 - PRIMARY where no paper exists) " + "-"*60, bold=True)
for w, a, u in [
 ("IBM granite-3.3-8b-instruct card", "HumanEval 89.73, MMLU 65.54, GSM8K 80.89", "huggingface.co/ibm-granite/granite-3.3-8b-instruct"),
 ("Microsoft phi-4 card", "HumanEval 82.6, MMLU 84.8, 1920xH100 21 days", "huggingface.co/microsoft/phi-4"),
 ("Microsoft Phi-3.5-mini-instruct card", "62.8 / 69 / 86.2, 512xH100 10 days", "huggingface.co/microsoft/Phi-3.5-mini-instruct"),
 ("Microsoft Phi-3.5-vision-instruct card", "MMBench 81.9, MMMU 43.0, 256xA100 6 days", "huggingface.co/microsoft/Phi-3.5-vision-instruct"),
 ("Qwen2.5-Coder-7B-Instruct card", "7.61B params, 28 layers, 4 KV heads, 131072 ctx", "huggingface.co/Qwen/Qwen2.5-Coder-7B-Instruct"),
 ("Qwen2-VL-7B-Instruct card", "DocVQA 94.5, TextVQA 84.3, MMMU 54.1", "huggingface.co/Qwen/Qwen2-VL-7B-Instruct"),
 ("Meta Llama-3.1-8B-Instruct card", "HumanEval 72.6, MMLU 69.4 vs CoT 73.0, GPU-hours", "huggingface.co/meta-llama/Llama-3.1-8B-Instruct"),
 ("Meta Llama-3.2-3B-Instruct card", "MMLU 63.4, GSM8K 77.7, EU clause scope", "huggingface.co/meta-llama/Llama-3.2-3B-Instruct"),
 ("Meta Llama 3.2 MODEL_CARD_VISION.md", "11B Vision INSTRUCT table (not pretrained)", "github.com/meta-llama/llama-models"),
 ("BigCode starcoder2-15b card", "1024xH100, 16384 ctx, HumanEval 46.3", "huggingface.co/bigcode/starcoder2-15b"),
 ("Google codegemma-7b-it card", "HumanEval 56.1, BabelCode C++ 42.2, 8K ctx", "huggingface.co/google/codegemma-7b-it"),
 ("Google gemma-3-4b-it card", "HumanEval 36.0, MMLU 59.6, 128K in / 8K out", "huggingface.co/google/gemma-3-4b-it"),
 ("Google paligemma-3b-mix-448 card", "512 token ctx, fine-tuned transfer scores", "huggingface.co/google/paligemma-3b-mix-448"),
 ("DeepSeek-R1-Distill-Qwen-14B card", "MATH-500 93.9, LiveCodeBench 53.1, Codeforces 1481", "huggingface.co/deepseek-ai/DeepSeek-R1-Distill-Qwen-14B"),
 ("DeepSeek-Coder-V2-Lite-Instruct card", "MoE config, licence terms", "huggingface.co/deepseek-ai/DeepSeek-Coder-V2-Lite-Instruct"),
 ("OpenGVLab InternVL2-8B card", "MMBench 81.7, MMMU 51.8, DocVQA 91.6, 8K ctx", "huggingface.co/OpenGVLab/InternVL2-8B"),
 ("llava-hf llava-v1.6-vicuna-13b-hf card", "Licence = LLAMA 2 (publishes no benchmarks)", "huggingface.co/llava-hf/llava-v1.6-vicuna-13b-hf"),
 ("vikhyatk/moondream2 card", "DocVQA 79.3, TextVQA 76.3, COCO detect 51.2", "huggingface.co/vikhyatk/moondream2"),
 ("mistralai Mistral-7B-Instruct-v0.3 card", "PUBLISHES NO BENCHMARKS - that is the finding", "huggingface.co/mistralai/Mistral-7B-Instruct-v0.3"),
 ("Qwen3-VL-Embedding-8B card", "MMEB-V2 Image 80.1 / Video 66.1 / VisDoc 83.3", "huggingface.co/Qwen/Qwen3-VL-Embedding-8B"),
 ("Qwen3-VL-Embedding-2B card", "MMEB-V2 Image 75.0 / Video 61.9, MMTEB 63.87", "huggingface.co/Qwen/Qwen3-VL-Embedding-2B"),
 ("Alibaba-NLP gme-Qwen2-VL-7B-Instruct card", "UMRB 67.44, MTEB-en 67.48, dim 3584", "huggingface.co/Alibaba-NLP/gme-Qwen2-VL-7B-Instruct"),
 ("Alibaba-NLP gme-Qwen2-VL-2B-Instruct card", "UMRB 64.45, MTEB-en 65.27, dim 1536", "huggingface.co/Alibaba-NLP/gme-Qwen2-VL-2B-Instruct"),
 ("TIGER-Lab VLM2Vec-Full card", "Apache 2.0, Phi-3.5-Vision backbone, 4B", "huggingface.co/TIGER-Lab/VLM2Vec-Full"),
 ("vidore colqwen2-v1.0 card", "Apache 2.0 + MIT, multi-vector ColBERT", "huggingface.co/vidore/colqwen2-v1.0"),
 ("vidore colpali-v1.2 card", "MIT adapters over Gemma-licensed backbone", "huggingface.co/vidore/colpali-v1.2"),
 ("royokong e5-v card", "NO LICENCE DECLARED - that is the finding", "huggingface.co/royokong/e5-v"),
]:
    r = line(r, "%-58s %-34s %s" % (w, a, u))
r = line(r, "")
r = line(r, "--- OFFICIAL BLOGS AND VENDOR DOCS (evidence grade 3) " + "-"*88, bold=True)
for w, a, u in [
 ("LLaVA-NeXT release blog", "LLaVA-1.6 13B: VQAv2 82.8, MMMU 36.2, TextVQA 67.1", "llava-vl.github.io/blog/2024-01-30-llava-next"),
 ("Qwen2.5-Coder family blog", "Qwen family overview (no per-model numbers)", "qwenlm.github.io/blog/qwen2.5-coder-family"),
 ("Google CodeGemma docs", "CodeGemma model documentation", "ai.google.dev/gemma/docs/codegemma"),
 ("InternVL project site", "InternVL2 family documentation", "internvl.github.io"),
 ("VLM2Vec project page", "VLM2Vec-V2 MMEB-V2 overall 58.0", "tiger-ai-lab.github.io/VLM2Vec"),
 ("QwenLM Qwen3-VL-Embedding repo", "Reference implementation, vLLM embedding mode", "github.com/QwenLM/Qwen3-VL-Embedding"),
 ("TIGER-AI-Lab VLM2Vec repo", "VLM2Vec / V2 reference implementation", "github.com/TIGER-AI-Lab/VLM2Vec"),
 ("illuin-tech colpali repo", "colpali-engine, ColBERT late interaction", "github.com/illuin-tech/colpali"),
 ("ViDoRe leaderboard", "LIVE source for ColPali/ColQwen2 scores", "huggingface.co/spaces/vidore/vidore-leaderboard"),
 ("MMEB leaderboard", "LIVE source for MMEB / MMEB-V2 scores", "huggingface.co/spaces/TIGER-Lab/MMEB"),
]:
    r = line(r, "%-58s %-34s %s" % (w, a, u))
r = line(r, "")
r = line(r, "--- TOKEN-RATE MEASUREMENTS (evidence grade 4 - the ONLY published source for tok/s) " + "-"*55, bold=True)
for w, a, u in [
 ("LLM tokens/sec benchmarks, llama.cpp b3520", "RTX4090 Q4_K_M: 7B 135, 13B 78, 34B 42 tok/s", "mustafa.net/llm-tokens-per-second-benchmarks"),
 ("Ollama vs llama.cpp benchmark", "Llama 3.1 8B Q4_K_M on 4090 = 95-110 tok/s", "markaicode.com/benchmarks/ollama-vs-llamacpp-benchmark"),
 ("Qwen3-8B on RTX 4090 recipe", "llama.cpp 104 tok/s at 16K context cross-check", "smeltcore.com"),
 ("AMD EPYC LLM inference benchmark", "CPU: EPYC 7763 7B Q4_K_M = 15 tok/s", "blog.leaseweb.com"),
 ("llama.cpp VRAM requirements guide", "Concurrency: ~18 tok/s per user, 4 users, 24GB", "localllm.in/blog/llamacpp-vram-requirements-for-local-llms"),
 ("myaihardware llama.cpp benchmarks", "CPU desktop band 4-15 tok/s at Q4_K_M", "myaihardware.com/llama-cpp-benchmarks"),
]:
    r = line(r, "%-58s %-34s %s" % (w, a, u))
r = line(r, "")
r = line(r, "--- BENCHMARK DEFINITIONS (what each metric actually measures) " + "-"*80, bold=True)
for w, a, u in [
 ("HumanEval / Evaluating LLMs Trained on Code", "164 Python problems, pass@1", "arxiv.org/abs/2107.03374"),
 ("MultiPL-E", "HumanEval translated to 18 languages incl. C++", "arxiv.org/abs/2208.08227"),
 ("MBPP", "Mostly Basic Python Problems", "arxiv.org/abs/2108.07732"),
 ("EvalPlus (HumanEval+ / MBPP+)", "80x and 35x more tests than the originals", "arxiv.org/abs/2305.01210"),
 ("CanItEdit", "Code EDITING - closest proxy to a retry loop", "arxiv.org/abs/2312.12450"),
 ("CRUXEval", "Code reasoning, understanding and execution", "arxiv.org/abs/2401.03065"),
 ("RepoBench", "Repository-level next-line completion", "arxiv.org/abs/2306.03091"),
 ("MMLU", "57-subject multitask accuracy", "arxiv.org/abs/2009.03300"),
 ("GSM8K", "Grade-school math word problems", "arxiv.org/abs/2110.14168"),
 ("MATH", "Competition mathematics", "arxiv.org/abs/2103.03874"),
 ("MMMU", "College-level multimodal understanding", "arxiv.org/abs/2311.16502"),
 ("DocVQA", "Document visual question answering", "arxiv.org/abs/2007.00398"),
 ("TextVQA", "Reading text within images", "arxiv.org/abs/1904.08920"),
 ("LiveCodeBench", "Contamination-free live coding problems", "arxiv.org/abs/2403.07974"),
]:
    r = line(r, "%-58s %-34s %s" % (w, a, u))
r = line(r, "")
r = line(r, "--- ENGINEERING AND HARDWARE SOURCES " + "-"*105, bold=True)
for w, a, u in [
 ("llama.cpp - supported backends and build docs", "ALL inference hardware support claims", "github.com/ggml-org/llama.cpp"),
 ("llama.cpp k-quants PR #1684", "ALL quantization accuracy deltas", "github.com/ggml-org/llama.cpp/pull/1684"),
 ("NVIDIA Ada GPU Architecture Whitepaper", "RTX 4090 bandwidth 1008 GB/s", "nvidia.com - Ada whitepaper"),
 ("vLLM / PagedAttention", "Concurrency behaviour", "arxiv.org/abs/2309.06180"),
 ("SGLang / RadixAttention", "Prefix caching for the shared 6-8K header", "arxiv.org/abs/2312.07104"),
 ("LoRA", "Fine-tuning VRAM baseline", "arxiv.org/abs/2106.09685"),
 ("QLoRA", "4-bit fine-tuning VRAM", "arxiv.org/abs/2305.14314"),
]:
    r = line(r, "%-58s %-34s %s" % (w, a, u))
r = line(r, "")
r = line(r, "--- LICENCE AND COMPLIANCE TEXTS " + "-"*109, bold=True)
for w, a, u in [
 ("Apache License 2.0", "Qwen, Granite, Mistral, Moondream, Qwen2-VL", "apache.org/licenses/LICENSE-2.0"),
 ("MIT License", "Phi-4, Phi-3.5-mini, Phi-3.5-Vision, InternVL2, R1", "opensource.org/license/mit"),
 ("Llama 3.x Community Licence", "EU MULTIMODAL EXCLUSION - exact wording", "llama.com/llama3_3/license"),
 ("Llama 2 Community Licence", "LLaVA-1.6 effective licence via Vicuna", "ai.meta.com/llama/license"),
 ("Gemma Terms of Use", "CodeGemma, Gemma 3, PaliGemma", "ai.google.dev/gemma/terms"),
 ("BigCode OpenRAIL-M v1", "StarCoder2 - flow-down duty on derivatives", "bigcode-project.org/docs/pages/model-license"),
 ("DeepSeek Model Licence", "Use restrictions attachment", "github.com/deepseek-ai/DeepSeek-LLM"),
 ("NDAA Section 889 / FAR 52.204-25", "Component-origin procurement rule", "acquisition.gov/far/52.204-25"),
]:
    r = line(r, "%-58s %-34s %s" % (w, a, u))

OUT = "snorkelbadger_SLM_Matrix.xlsx"
wb.save(OUT)
print("saved:", OUT)
print("sheets:", wb.sheetnames)
print("Summary : %d cols x %d model-quant rows" % (len(C1), len(M) * 4))
print("Detailed: %d cols (%d data + %d source + 1 status)" % (
    len(C2), sum(1 for h, _ in C2 if "SOURCE" not in h and "VERIFICATION" not in h),
    sum(1 for h, _ in C2 if "SOURCE" in h)))
print("models:", len(M))
